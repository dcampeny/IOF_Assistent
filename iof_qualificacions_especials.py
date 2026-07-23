# -*- coding: utf-8 -*-
"""
iof_qualificacions_especials.py

Eina per carregar cartografia de referència de qualificacions especials
(espais naturals protegits) i calcular quina superfície de cada
tipologia forestal hi queda afectada, amb el resultat llest per
traspassar al formulari oficial del PTGMF (secció 1.3 "Qualificacions
especials i afectacions").

Fonts de dades (investigades i verificades juliol 2026 — vegeu CLAUDE.md
per al detall complet de la investigació):

- WFS `sig.gencat.cat/ows/ESPAIS_NATURALS/wfs`: ENPE (01), PEIN (03) i
  PEIN-PE (02) (mateixa capa, distingides pel camp PLANIF), Xarxa
  Natura 2000 -> LIC (13) / ZEPA (12) / LIC-ZEPA (11) (mateixa capa,
  distingides pels camps LIC_ZEC/ZEPA), UP (09).
  RF (04) s'assimila a "Reserva Natural", ja inclosa dins la capa ENPE
  (camps CODI_RNI/CODI_RNP/CODI_RNFS) -- no necessita font pròpia.
- WFS `sig.gencat.cat/ows/FAUNA/wfs`: FAUNA (06), capa FAUNA_AIFF_PUBLICA.
- Descàrrega directa de fitxer SHP (agricultura.gencat.cat): PPP (14),
  ZAU (10).
- WMS + GetFeatureInfo (dtes.gencat.cat, Mapa Urbanístic de Catalunya):
  LU (08) -- només es pot consultar per punt (WFS rebutjat pel servidor),
  filtrant pels codis CODI_QUAL_MUC que comencen per "N2" (protecció
  establerta pel propi planejament urbanístic; "N3" ja queda cobert per
  ENPE/PEIN/Xarxa Natura 2000 i NO es compta, per no duplicar superfície).
- BS (15, Boscos Singulars): SENSE CARTOGRAFIA DISPONIBLE -- no
  automatitzable, es mostra com a nota manual a l'informe.

Totes les peticions a sig.gencat.cat necessiten una capçalera User-Agent
(sense ella, el servidor retorna 403).

IMPORTANT -- abast de les proves fetes: el carregament de capes (WFS,
descàrrega SHP, WMS) s'ha verificat en directe amb qgis-mcp. El càlcul
de superposicions (native:clip/native:intersection) i el mostreig per
punt de LU (GetFeatureInfo) shan escrit seguint el mateix patró ja
verificat a la sessió d'investigació, pero cal validar-ho de nou en
directe abans de donar-ho per definitiu -- vegeu CLAUDE.md.
"""
import os
import re
import json
import urllib.request
import urllib.parse
import zipfile
import tempfile

from qgis.PyQt.QtWidgets import QMessageBox, QApplication, QFileDialog
from qgis.PyQt.QtCore import Qt

from qgis.core import (
    QgsProject, QgsVectorLayer, QgsRasterLayer, QgsField, QgsFeature,
    QgsVectorFileWriter, QgsCoordinateTransformContext, QgsRectangle,
    QgsProviderRegistry, QgsGeometry, QgsPointXY,
    QgsMessageLog, Qgis,
)
from qgis.PyQt.QtCore import QVariant

from .iof_utils import get_layer as _get_layer
from .iof_qualificacions_creuament import (
    cerca_correspondencia_pdf as _cerca_correspondencia_pdf,
    CREUAMENT_PDF as _CREUAMENT_PDF,
)


def _log_error(context, exc):
    """Registra una excepció al log de missatges de QGIS (pestanya "IOF
    Assistent"), amb traça completa -- perquè un error durant la
    càrrega d'una qualificació especial no quedi silenciat sense cap
    rastre (abans només s'afegia el codi a `errors`, sense cap detall
    del què havia fallat ni per què)."""
    import traceback
    QgsMessageLog.logMessage(
        context + ": " + repr(exc) + "\n" + traceback.format_exc(),
        "IOF Assistent", Qgis.MessageLevel.Warning,
    )

# Noms de la capa PEIN (NOM_PEIN) que corresponen a RF (04) -- RF NO ve
# d'ENPE (correcció de l'usuari: l'assimilació anterior a ENPE era un
# error), es creua directament amb PEIN, identificat pels 3 noms
# concrets ja validats manualment.
_NOMS_PEIN_QUE_SON_RF = set(_CREUAMENT_PDF.get("RF@PEIN", {}).keys())


USER_AGENT = {"User-Agent": "Mozilla/5.0 (QGIS IOF_Assistent)"}

LAYERS_UNITATS = ["IOF_Unitats_Actuacio", "IOF_Rodals"]
LAYER_AMBIT_NAME = "Àmbit IOF"
GRUP_NOM = "Qualificacions especials"
GRUP_ALTRES_NOM = "Altres qualificacions especials"

WFS_ESPAIS_NATURALS = "https://sig.gencat.cat/ows/ESPAIS_NATURALS/wfs"
WFS_FAUNA = "https://sig.gencat.cat/ows/FAUNA/wfs"
URL_PPP = "https://www.gencat.cat/agricultura/sig/bases/perprot.zip"
URL_ZAU = "https://www.gencat.cat/agricultura/sig/bases/zaus.zip"
WMS_MUC = "https://dtes.gencat.cat/webmap/MUC/service.svc/get"

# Mapa de risc d'incendi tipus de Catalunya (Centre de la Propietat
# Forestal / GRAF-Bombers) -- descàrrega directa SHP, camp RISC amb 4
# nivells (Baix/Moderat/Alt/Molt alt).
URL_RISC_INCENDI_TIPUS = (
    "https://cpf.gencat.cat/web/.content/or_organismes/or04_centre_propietat_forestal/"
    "03_linies_actuacio/linies_dactuacio/transferencia_de_coneixement/orgest/"
    "servidor_i_desc_rrega_de_cartografia/mapa_de_risc_d_incendi_tipus_a_catalunya/Risc_Inc.zip"
)

# Noms complets per a la llegenda del mapa (nom sencer + sigla entre
# parèntesi) -- els codis curts es mantenen intactes com a nom intern
# de cada capa dins IOF_Qualificacions.gpkg (necessaris perquè
# "Exportar qualificacions especials" les trobi per aquest nom exacte
# amb _carrega_capa_font()); només es tradueixen en el moment de
# mostrar-les al mapa.
NOMS_LLEGENDA = {
    "ENPE": "Espai Natural de Protecció Especial (ENPE)",
    "PEIN": "Pla d'Espais d'Interès Natural (PEIN)",
    "Xarxa_Natura_2000": "Xarxa Natura 2000",
    "UP": "Espai Catalogat d'Utilitat Pública (UP)",
    "FAUNA": "Àrea de presència de fauna protegida",
    "PPP": "Perímetre de Protecció Prioritària (PPP)",
    "ZAU": "Zona d'Actuació Urgent (ZAU)",
}

# Noms de llegenda pels SUBTIPUS de PEIN i Xarxa Natura 2000 (es
# mostren com a capes separades dins el mateix grup): mateix criteri
# que NOMS_LLEGENDA -- nom sencer + sigles entre parèntesi -- perquè
# no quedin amb només la sigla a la llegenda del mapa.
NOMS_LLEGENDA_SUBTIPUS = {
    "PEIN-PE": "Pla d'Espais d'Interès Natural amb Pla especial (PEIN-PE)",
    "PEIN": "Pla d'Espais d'Interès Natural (PEIN)",
    "LIC-ZEPA": "Lloc d'Importància Comunitària i Zona d'Especial Protecció per a les Aus (LIC-ZEPA)",
    "ZEPA": "Zona d'Especial Protecció per a les Aus (ZEPA)",
    "LIC": "Lloc d'Importància Comunitària (LIC)",
}

# Noms de qualificació per a l'INFORME D'EXPORTACIÓ (columna "Qualificació"
# de l'Excel/text) -- diferent del nom de la llegenda del mapa: aquí
# l'usuari ha demanat un text més senzill, sense repetir informació
# entre parèntesi quan el codi ja n'hi és, o amb la sigla desenvolupada
# quan calgui llegir-ho sol (p. ex. a "Observacions").
NOMS_QUALIFICACIO_INFORME = {
    "ENPE": "ENPE",
    "PEIN": "PEIN",
    "PEIN-PE": "PEIN-PE",
    "RF": "RF",
    "LIC": "LIC",
    "ZEPA": "ZEPA",
    "LIC-ZEPA": "LIC-ZEPA",
    "UP": "UP",
    "FAUNA": "FAUNA",
    "ZAU": "ZAU",
    "LU": "LU",
}

# Citació de la font de cada capa (Departament/organisme + servei
# d'origen) -- per a les metadades de cada capa i per al resum final
# mostrat a l'usuari, de cara a citar-ho correctament en preparar el
# plànol imprès de l'IOF.
FONTS_CITACIO = {
    "ENPE": "Generalitat de Catalunya. Servei WFS ESPAIS_NATURALS (sig.gencat.cat)",
    "PEIN": "Generalitat de Catalunya. Servei WFS ESPAIS_NATURALS (sig.gencat.cat)",
    "Xarxa_Natura_2000": "Generalitat de Catalunya. Servei WFS ESPAIS_NATURALS (sig.gencat.cat)",
    "UP": "Generalitat de Catalunya. Servei WFS ESPAIS_NATURALS (sig.gencat.cat)",
    "FAUNA": "Generalitat de Catalunya. Servei WFS FAUNA (sig.gencat.cat)",
    "PPP": "Departament d'Agricultura, Ramaderia, Pesca i Alimentació. Generalitat de Catalunya (agricultura.gencat.cat)",
    "ZAU": "Departament d'Agricultura, Ramaderia, Pesca i Alimentació. Generalitat de Catalunya (agricultura.gencat.cat)",
}


def _aplica_metadata(layer, nom_codi):
    """Desa la font/citació de la capa a les seves metadades (Propietats
    de la capa → Metadades → Resum), perquè quedi enregistrada la
    procedència de les dades dins el propi projecte."""
    from qgis.core import QgsLayerMetadata
    citacio = FONTS_CITACIO.get(nom_codi)
    if not citacio:
        return
    meta = layer.metadata()
    meta.setAbstract(
        "Font: " + citacio + ". Descarregat automàticament pel "
        "complement IOF Assistent (eina Qualificacions especials)."
    )
    meta.setRights([citacio])
    layer.setMetadata(meta)



# ---------------------------------------------------------------------------
# Utilitats compartides
# ---------------------------------------------------------------------------

def _troba_unitats_layer():
    for name in LAYERS_UNITATS:
        lyr = _get_layer(name)
        if lyr is not None:
            return lyr
    return None


def _troba_nom_iof():
    """Retorna el nom de l'IOF (camp `nom_finca` de la capa IOF_Finques,
    unint els noms de totes les finques si n'hi ha més d'una), per fer-
    lo servir com a part del nom de fitxer de l'informe. Retorna "" si
    no es pot determinar."""
    lyr = _get_layer("IOF_Finques")
    if lyr is None or lyr.featureCount() == 0:
        return ""
    if "nom_finca" not in lyr.fields().names():
        return ""
    noms = []
    for feat in lyr.getFeatures():
        v = (feat["nom_finca"] or "").strip()
        if v and v not in noms:
            noms.append(v)
    return " i ".join(noms)


def _format_ha_ca(value, decimals=2):
    """Formata un número amb coma decimal (convenció catalana) en lloc
    del punt per defecte de Python -- per construir el text dels
    AVISOS (QMessageBox), on els números es concatenen manualment com
    a text. No afecta les cel·les numèriques de l'Excel: allà Excel ja
    aplica la configuració regional de l'usuari (per això ja es veuen
    amb coma al full de càlcul)."""
    return format(value, "." + str(decimals) + "f").replace(".", ",")


def _troba_camp_codi(layer):
    if "codi_ua" in layer.fields().names():
        return "codi_ua"
    if "codi_rodal" in layer.fields().names():
        return "codi_rodal"
    return None


def _wfs_uri(url, typename, crs="EPSG:25831"):
    return (
        "pagingEnabled='true' restrictToRequestBBOX='1' srsname='" + crs +
        "' typename='" + typename + "' url='" + url + "' version='2.0.0'"
    )


def _descarrega_shp_zip(url, prefix):
    """Descarrega un ZIP amb un SHP (amb capçalera User-Agent, necessària
    per a agricultura/gencat.cat), el descomprimeix i retorna el path
    del fitxer .shp."""
    if not url.lower().startswith("https://"):
        raise ValueError("URL no permesa (només HTTPS): " + url)
    req = urllib.request.Request(url, headers=USER_AGENT)
    tmp_dir = tempfile.mkdtemp(prefix="iof_qe_" + prefix + "_")
    zip_path = os.path.join(tmp_dir, prefix + ".zip")
    # URL validada com a HTTPS just abans; mai ve de l'usuari (constant fixa)
    with urllib.request.urlopen(req, timeout=30) as resp:  # nosec B310
        data = resp.read()
    with open(zip_path, "wb") as f:
        f.write(data)
    with zipfile.ZipFile(zip_path) as z:
        names = z.namelist()
        z.extractall(tmp_dir)
    shp_candidates = [n for n in names if n.lower().endswith(".shp")]
    if not shp_candidates:
        raise RuntimeError("El ZIP descarregat no conté cap fitxer .shp")
    return os.path.join(tmp_dir, shp_candidates[0])


def _grup_qualificacions(proj):
    root = proj.layerTreeRoot()
    grup = root.findGroup(GRUP_NOM)
    if grup is None:
        grup = root.insertGroup(0, GRUP_NOM)
    return grup


def _grup_altres_qualificacions(proj):
    """Segon grup, germà de "Qualificacions especials" (s'insereix just
    a sota): hi van totes les qualificacions especials que queden per
    sota de la primera en jerarquia (PEIN-PE, PEIN, LIC-ZEPA, ZEPA,
    LIC) més PPP, UP, FAUNA i ZAU. Es carreguen igual (GeoPackage,
    estil, etiqueta) però amb la visibilitat desmarcada per defecte
    (`setItemVisibilityChecked(False)`, aplicat a _mostra_amb_vores_reals
    amb visible=False): no es dibuixen al mapa per no saturar-lo, però
    hi són disponibles per activar-les manualment i per a qualsevol
    consulta/exportació posterior."""
    root = proj.layerTreeRoot()
    grup = root.findGroup(GRUP_ALTRES_NOM)
    if grup is None:
        grup_principal = root.findGroup(GRUP_NOM)
        index = (root.children().index(grup_principal) + 1) if grup_principal else 0
        grup = root.insertGroup(index, GRUP_ALTRES_NOM)
    return grup


def _troba_extent_topografic(iface):
    """
    Cerca el grup de capes "Topogràfic territorial N" (creat en carregar
    "Referencial topogràfic territorial vectorial", dins
    iof_gestor_topografia_dialog.py) i en retorna l'extensió del
    rectangle de descàrrega original -- l'àrea que s'ha d'utilitzar per
    a les qualificacions especials, perquè coincideixi exactament amb
    la mateixa zona ja delimitada per l'usuari amb l'Open ICGC.

    IMPORTANT: no es fa la unió (bounding box combinat) de totes les
    subcapes -- les subcapes de punts (sufixos "_n"/"_p": nodes,
    topònims, etc.) solen tenir una extensió lleugerament més petita o
    descentrada respecte al rectangle real (cap punt necessàriament
    toca la vora exacta), i incloure-les inflaria o desplaçaria el
    resultat. En canvi, la gran majoria de subcapes de línies/polígons
    (que cobreixen tota l'àrea de manera contínua) coincideixen EXACTAMENT
    en el mateix rectangle -- aquest és el que es fa servir, triant
    l'extensió més freqüent (moda) entre totes les subcapes.

    Si n'hi ha diversos grups "Topogràfic territorial N" (s'ha carregat
    més d'un cop), es fa servir el més recent (número més alt). Retorna
    None si no se'n troba cap.
    """
    root = QgsProject.instance().layerTreeRoot()
    grup_pare = root.findGroup("Cartografia de referència")
    if grup_pare is None:
        return None

    grups_topo = [
        child for child in grup_pare.children()
        if hasattr(child, "name") and child.name().startswith("Topogràfic territorial")
    ]
    if not grups_topo:
        return None

    def _numero(g):
        try:
            return int(g.name().replace("Topogràfic territorial", "").strip())
        except ValueError:
            return 0

    grups_topo.sort(key=_numero, reverse=True)
    grup_escollit = grups_topo[0]

    comptador = {}
    for layer_node in grup_escollit.findLayers():
        lyr = layer_node.layer()
        if lyr is None:
            continue
        lyr_extent = lyr.extent()
        if lyr_extent.isEmpty():
            continue
        # S'arrodoneix a 0,1 m per considerar coincidents petites
        # diferències numèriques irrellevants entre capes.
        clau = (
            round(lyr_extent.xMinimum(), 1), round(lyr_extent.yMinimum(), 1),
            round(lyr_extent.xMaximum(), 1), round(lyr_extent.yMaximum(), 1),
        )
        comptador[clau] = comptador.get(clau, 0) + 1

    if not comptador:
        return None

    clau_moda = max(comptador.items(), key=lambda kv: kv[1])[0]
    return QgsRectangle(clau_moda[0], clau_moda[1], clau_moda[2], clau_moda[3])


def _retalla_per_extent(layer, extent, clip=True):
    """Selecciona les entitats dins l'extensió donada amb
    native:extractbyextent. Si clip=True (per defecte), RETALLA-LES
    exactament a la vora del rectangle; si clip=False, es queda amb
    la geometria SENCERA de cada entitat seleccionada (útil com a
    filtre previ ràpid per reduir volum, sense perdre'n la forma).

    NOTA: es va provar l'alternativa (CLIP=False, mantenint la
    geometria sencera) per evitar que un espai gran es tallés amb una
    vora artificial. Però amb CLIP=False, espais com "Serres de
    Montnegre-el Corredor" es dibuixaven estenent-se molt més enllà de
    la zona rellevant, donant un aspecte desordenat al mapa -- l'usuari
    ho va veure en pantalla i va confirmar explícitament que prefereix
    tornar a CLIP=True (retallar exactament al rectangle), acceptant
    que pugui tornar a aparèixer alguna vora recta allà on es talla un
    espai gran.

    Aquest canvi només afecta la VISUALITZACIÓ i el GeoPackage
    (`IOF_Qualificacions.gpkg`) -- no afecta la correcció del càlcul a
    "Exportar qualificacions especials", ja que aquest sempre retalla
    de nou contra l'Àmbit IOF (sempre més petit que aquest rectangle),
    així que el resultat final és idèntic independentment de si les
    dades del GeoPackage estan retallades al rectangle o no."""
    import processing
    res = processing.run("native:extractbyextent", {
        "INPUT": layer, "EXTENT": extent, "CLIP": clip,
        "OUTPUT": "TEMPORARY_OUTPUT",
    })
    return res["OUTPUT"]


def _selecciona_per_interseccio_real(layer, ambit_lyr):
    """Selecciona (SENSE retallar la geometria) les entitats de LAYER
    que intersequen realment el polígon de AMBIT_LYR -- predicat
    "intersecta" (índex 0) via native:extractbylocation. A diferència
    de native:clip, es queda amb la forma SENCERA de cada entitat
    seleccionada, no només el tros que cauria dins l'àmbit: criteri
    per a FAUNA/UP/ZAU -- si una entitat toca realment l'Àmbit IOF,
    es vol veure la seva forma global (retallada després al
    rectangle del referencial topogràfic, com ENPE/PEIN), no només
    el fragment que cau dins la finca."""
    import processing
    res = processing.run("native:extractbylocation", {
        "INPUT": layer, "PREDICATE": [0], "INTERSECT": ambit_lyr,
        "OUTPUT": "TEMPORARY_OUTPUT",
    })
    return res["OUTPUT"]


def _escriu_capa_gpkg(layer, gpkg_path, layer_name, es_primera):
    """Escriu una capa dins un GeoPackage amb diverses capes: la primera
    crea (o sobreescriu) el fitxer; les següents s'hi afegeixen com a
    capa nova sense esborrar les ja existents.

    Com que dins el mateix procés es van recarregant capes d'aquest
    mateix fitxer per mostrar-les al mapa (`lyr_carregada = QgsVectorLayer(
    gpkg_path + "|layername=...")`) abans d'escriure la capa següent,
    a Windows això pot deixar una connexió GDAL/SQLite oberta que
    bloqueja momentàniament l'escriptura -- provat en directe: un error
    de "PermissionError" en intentar reescriure/eliminar un GeoPackage
    amb una capa seva encara carregada. Es reintenta l'escriptura fins
    a 3 cops amb una pausa breu abans de donar l'error per definitiu."""
    import time
    options = QgsVectorFileWriter.SaveVectorOptions()
    options.driverName = "GPKG"
    options.layerName = layer_name
    options.actionOnExistingFile = (
        QgsVectorFileWriter.ActionOnExistingFile.CreateOrOverwriteFile
        if es_primera else
        QgsVectorFileWriter.ActionOnExistingFile.CreateOrOverwriteLayer
    )
    intents = 3
    for intent in range(1, intents + 1):
        error, msg, _, _ = QgsVectorFileWriter.writeAsVectorFormatV3(
            layer, gpkg_path, QgsCoordinateTransformContext(), options
        )
        if error == QgsVectorFileWriter.WriterError.NoError:
            return
        if intent < intents:
            time.sleep(0.4)
    raise RuntimeError("Error escrivint la capa " + layer_name + " al GeoPackage: " + msg)


# ---------------------------------------------------------------------------
# Botó 1: "Qualificacions especials afectades" -- carregar mapes de referència
# ---------------------------------------------------------------------------

def _wkt_vora_rectangle(extent):
    """Retorna el WKT de la vora (LINESTRING tancat) d'un QgsRectangle,
    per incrustar-lo dins una expressió de símbol."""
    return (
        "LINESTRING(" +
        str(extent.xMinimum()) + " " + str(extent.yMinimum()) + ", " +
        str(extent.xMaximum()) + " " + str(extent.yMinimum()) + ", " +
        str(extent.xMaximum()) + " " + str(extent.yMaximum()) + ", " +
        str(extent.xMinimum()) + " " + str(extent.yMaximum()) + ", " +
        str(extent.xMinimum()) + " " + str(extent.yMinimum()) + ")"
    )


def _crea_simbol_amb_halo(color_contorn, extent=None, estil_linia="dash",
                           gruix_linia=0.6, gruix_halo=1.0, tolerancia=0.5):
    """
    Crea un QgsFillSymbol de només contorn (sense farciment), amb un
    halo blanc per darrere de la línia de color -- mateixa tècnica que
    ja fa servir el complement per als límits de rodal/unitat
    d'actuació (`_sym_rodal()` a iof_format_dialog.py):
    QgsGeometryGeneratorSymbolLayer, amb dues capes de línia al mateix
    rendering pass (blanca més ampla primer, de color més fina
    després) -- així, si diverses qualificacions se superposen, cap
    línia queda il·legible enmig d'altres línies ni del fons del mapa.

    Si `extent` es dona, la geometria de la línia NO és tot el contorn
    del polígon (`boundary($geometry)`), sinó el contorn MENYS els
    trams que coincideixen amb la vora d'`extent` (una sola expressió
    QGIS, `difference(boundary($geometry), buffer(vora_rectangle,
    tolerancia))`) -- necessari perquè un polígon retallat amb
    `CLIP=True` a un rectangle no dibuixi aquesta vora de tall com si
    fos un límit real de l'espai. Tot en UNA sola capa (el mateix
    polígon, amb l'etiqueta si cal) -- NO es crea cap capa addicional.
    """
    from qgis.core import (
        QgsFillSymbol, QgsSimpleFillSymbolLayer, QgsGeometryGeneratorSymbolLayer,
        QgsSimpleLineSymbolLayer, QgsLineSymbol,
    )
    from qgis.PyQt.QtGui import QColor
    from qgis.PyQt.QtCore import Qt

    pen_styles = {
        "solid": Qt.PenStyle.SolidLine, "dash": Qt.PenStyle.DashLine,
        "dot": Qt.PenStyle.DotLine, "dash dot": Qt.PenStyle.DashDotLine,
        "dash dot dot": Qt.PenStyle.DashDotDotLine,
    }

    if extent is not None:
        geom_modifier = (
            "difference(boundary($geometry), buffer(geom_from_wkt('" +
            _wkt_vora_rectangle(extent) + "'), " + str(tolerancia) + "))"
        )
    else:
        geom_modifier = "boundary($geometry)"

    fill = QgsSimpleFillSymbolLayer.create({"color": "0,0,0,0", "outline_style": "no"})
    fill.setRenderingPass(0)

    # Halo blanc (per darrere)
    line_gen_halo = QgsGeometryGeneratorSymbolLayer.create({
        "geometryModifier": geom_modifier, "SymbolType": "Line",
    })
    lyr_halo = QgsSimpleLineSymbolLayer()
    lyr_halo.setColor(QColor(255, 255, 255))
    lyr_halo.setWidth(gruix_halo)
    lyr_halo.setPenStyle(Qt.PenStyle.SolidLine)
    sym_halo = QgsLineSymbol()
    sym_halo.deleteSymbolLayer(0)
    sym_halo.appendSymbolLayer(lyr_halo)
    line_gen_halo.setSubSymbol(sym_halo)
    line_gen_halo.setRenderingPass(2)

    # Línia de color (per sobre, mateix rendering pass -- s'afegeix després)
    line_gen_color = QgsGeometryGeneratorSymbolLayer.create({
        "geometryModifier": geom_modifier, "SymbolType": "Line",
    })
    lyr_color = QgsSimpleLineSymbolLayer()
    parts = [int(x) for x in color_contorn.split(",")]
    lyr_color.setColor(QColor(*parts))
    lyr_color.setWidth(gruix_linia)
    lyr_color.setPenStyle(pen_styles.get(estil_linia, Qt.PenStyle.DashLine))
    sym_color = QgsLineSymbol()
    sym_color.deleteSymbolLayer(0)
    sym_color.appendSymbolLayer(lyr_color)
    line_gen_color.setSubSymbol(sym_color)
    line_gen_color.setRenderingPass(2)

    sym = QgsFillSymbol()
    sym.deleteSymbolLayer(0)
    sym.appendSymbolLayer(fill)
    sym.appendSymbolLayer(line_gen_halo)
    sym.appendSymbolLayer(line_gen_color)
    return sym


def _mostra_amb_vores_reals(proj, grup, polygon_layer, extent, color, estil_linia,
                             gruix_linia=0.6, camp_etiqueta=None, color_etiqueta=None,
                             color_halo_etiqueta="255,255,255,255", es_expressio_etiqueta=False,
                             visible=True):
    """
    Prepara i afegeix al mapa una qualificació especial evitant que la
    vora del rectangle de retall es dibuixi com si fos un límit real
    -- SENSE crear cap capa addicional: el mateix `polygon_layer` porta
    tant el símbol visible (contorn amb halo, calculat amb l'expressió
    `difference(boundary($geometry), buffer(vora_rectangle, ...))` via
    `_crea_simbol_amb_halo(..., extent=extent)`) com l'etiqueta
    (`Placement.Horizontal`, que fa servir sempre la geometria pròpia
    del polígon per centrar el text, no la del símbol calculat -- per
    això funciona igual de bé encara que el símbol dibuixi només línies).

    `visible=False` deixa la capa carregada i dins el grup, però amb
    la casella de visibilitat desmarcada (no es dibuixa al canvas) --
    es fa servir per a "Altres qualificacions especials", perquè
    quedin disponibles sense saturar el mapa per defecte.
    """
    from qgis.core import QgsSingleSymbolRenderer
    sym = _crea_simbol_amb_halo(color, extent=extent, estil_linia=estil_linia, gruix_linia=gruix_linia)
    polygon_layer.setRenderer(QgsSingleSymbolRenderer(sym))
    polygon_layer.triggerRepaint()
    if camp_etiqueta and color_etiqueta:
        _aplica_etiqueta(
            polygon_layer, camp_etiqueta, color_etiqueta,
            es_expressio=es_expressio_etiqueta, color_halo=color_halo_etiqueta,
        )
    proj.addMapLayer(polygon_layer, False)
    grup.addLayer(polygon_layer)
    if not visible:
        node = grup.findLayer(polygon_layer.id())
        if node:
            node.setItemVisibilityChecked(False)


def _aplica_etiqueta(layer, camp, color_text, es_expressio=False, color_halo="255,255,255,255"):
    """
    Aplica una etiqueta a una capa de polígons, seguint exactament la
    mateixa convenció ja establerta al complement per a la resta de
    capes (p. ex. `apply_infra_style()` a iof_format_dialog.py: Calibri
    9pt negreta, halo blanc 0.8) -- i el catàleg MTN25 adjuntat per
    l'usuari, que etiqueta els espais naturals amb "K040 Textos
    Negros" (text amb halo, no cap altre efecte).

    Si `es_expressio=True`, `camp` es tracta com una expressió QGIS
    (p. ex. per combinar diversos camps en una sola etiqueta) en lloc
    d'un simple nom de camp.

    `color_halo` (per defecte blanc): colors clars com el groc tenen
    molt poc contrast amb un halo blanc i costen de llegir -- es pot
    fer servir un halo més fosc (p. ex. gris) per a aquests casos.
    """
    from qgis.core import (
        QgsPalLayerSettings, QgsTextFormat, QgsTextBufferSettings,
        QgsVectorLayerSimpleLabeling,
    )
    from qgis.PyQt.QtGui import QFont, QColor as _QColor

    pal = QgsPalLayerSettings()
    pal.fieldName = camp
    pal.isExpression = es_expressio
    pal.enabled = True
    pal.placement = QgsPalLayerSettings.Placement.Horizontal

    fmt = QgsTextFormat()
    font = QFont("Calibri", 9)
    font.setBold(True)
    fmt.setFont(font)
    fmt.setSize(9)
    parts = [int(x) for x in color_text.split(",")]
    fmt.setColor(_QColor(*parts))

    buf = QgsTextBufferSettings()
    buf.setEnabled(True)
    buf.setSize(0.8)
    parts_halo = [int(x) for x in color_halo.split(",")]
    buf.setColor(_QColor(*parts_halo))
    fmt.setBuffer(buf)

    pal.setFormat(fmt)
    layer.setLabeling(QgsVectorLayerSimpleLabeling(pal))
    layer.setLabelsEnabled(True)


def _aplica_estil_nomes_limits(layer, color_contorn, estil_linia="dash", gruix="0.6"):
    """Aplica un estil de només contorn amb halo blanc (sense farciment)
    a una capa de polígons -- per a totes les qualificacions
    visualitzades. Estil per defecte "dash" (mai "solid"): el catàleg
    oficial MTN25 (IGN) no fa servir traç continu per a cap límit
    d'espai natural protegit ("Parque Nacional"=discontinu, "Parque
    Natural"=discontinu-puntejat)."""
    from qgis.core import QgsSingleSymbolRenderer
    symbol = _crea_simbol_amb_halo(color_contorn, estil_linia=estil_linia, gruix_linia=float(gruix))
    layer.setRenderer(QgsSingleSymbolRenderer(symbol))
    layer.triggerRepaint()


def _aplica_estil_pein(layer):
    """Estil per la capa PEIN, amb halo blanc -- una única simbologia
    per a PEIN i PEIN-PE junts (l'usuari ho ha demanat així: ja no es
    distingeixen visualment, encara que la distinció es manté al
    càlcul/informe de "Exportar qualificacions especials"). Mateix
    color que ENPE (#00A77E), diferenciat només per l'estil de línia."""
    _aplica_estil_nomes_limits(layer, "0,167,126,255", estil_linia="dash")


def _aplica_estil_natura2000(layer, camp_lic, camp_zepa):
    """Estil per la capa Xarxa Natura 2000, amb halo blanc, distingint
    LIC (blau, discontinu), ZEPA (blau cel, puntejat) i LIC-ZEPA (blau
    fosc, discontinu-doble puntejat) segons els camps LIC_ZEC/ZEPA.
    Sense traç continu, seguint la convenció del catàleg MTN25."""
    from qgis.core import QgsRuleBasedRenderer

    simbol_lic = _crea_simbol_amb_halo("63,81,181,255", estil_linia="dash")
    simbol_zepa = _crea_simbol_amb_halo("121,134,203,255", estil_linia="dot")
    simbol_liczepa = _crea_simbol_amb_halo("26,35,126,255", estil_linia="dash dot dot")

    es_lic = '"' + camp_lic + '" = \'Sí\''
    es_zepa = '"' + camp_zepa + '" = \'Sí\''
    arrel = QgsRuleBasedRenderer.Rule(None)
    arrel.appendChild(QgsRuleBasedRenderer.Rule(
        simbol_liczepa, 0, 0, "(" + es_lic + ") AND (" + es_zepa + ")",
        "Lloc d'Interès Comunitari i Zona d'Especial Protecció per a les Aus (LIC-ZEPA)"))
    arrel.appendChild(QgsRuleBasedRenderer.Rule(
        simbol_lic, 0, 0, "(" + es_lic + ") AND NOT (" + es_zepa + ")",
        "Lloc d'Interès Comunitari (LIC)"))
    arrel.appendChild(QgsRuleBasedRenderer.Rule(
        simbol_zepa, 0, 0, "NOT (" + es_lic + ") AND (" + es_zepa + ")",
        "Zona d'Especial Protecció per a les Aus (ZEPA)"))
    layer.setRenderer(QgsRuleBasedRenderer(arrel))
    layer.triggerRepaint()


def run_carregar_qualificacions(iface, parent=None):
    """
    Retalla totes les capes de qualificacions especials disponibles i
    les ajunta totes en un únic GeoPackage "IOF_Qualificacions.gpkg"
    (per a l'exportació/càlcul posterior), i les carrega al projecte
    en DOS grups:

    - "Qualificacions especials" (visible per defecte): ENPE,
      PEIN-PE, PEIN, LIC-ZEPA, ZEPA, LIC i FAUNA -- però només quan
      tenen una part pròpia i distintiva a mostrar. Per a ENPE/
      PEIN-PE/PEIN/LIC-ZEPA/ZEPA/LIC això vol dir: un cop retallada
      per la jerarquia (ENPE > PEIN-PE > PEIN > LIC-ZEPA > ZEPA >
      LIC), encara queda superfície pròpia, no coberta del tot per
      una qualificació de prioritat més alta.
    - "Altres qualificacions especials" (oculta per defecte, però
      disponible per activar manualment): PPP, UP i ZAU -- SEMPRE hi
      van, independentment de si toquen o no la finca. A més, quan
      una de ENPE/PEIN-PE/PEIN/LIC-ZEPA/ZEPA/LIC queda completament
      per sota d'una altra (superfície ja coberta del tot per una de
      prioritat més alta, per tant amb la forma retallada buida, no
      es pot veure), es mostra aquí la seva forma ORIGINAL sencera
      (sense retallar per jerarquia), per deixar constància que la
      finca també hi és afectada encara que no es vegi al grup
      principal.

    Regles de retall/visualització segons la qualificació:

    - ENPE, PEIN, Xarxa Natura 2000, PPP: es retallen i visualitzen a
      TOT el rectangle del Referencial topogràfic (només línies de
      contorn, sense farciment, per no tapar la resta de cartografia).
    - FAUNA, UP, ZAU: només es mostren si l'entitat toca REALMENT el
      polígon de l'Àmbit IOF (test d'intersecció exacte, no només el
      seu rectangle envolupant); si hi toca, es visualitza retallada
      a TOT el rectangle del Referencial topogràfic (com ENPE/PEIN),
      per veure'n la forma global i no només el fragment dins la
      finca.
    - LU: NO es visualitza al mapa (el càlcul real es fa per punt amb
      GetFeatureInfo a "Exportar qualificacions especials", no calen
      dades pròpies aquí).
    - BS: sense cartografia disponible, no es pot ni visualitzar ni
      exportar.

    Requereix haver carregat prèviament "Referencial topogràfic
    territorial vectorial" (Cartografia de referència) i tenir creat
    l'Àmbit IOF (Cadastre → Crear àmbit de l'IOF).
    """
    win = parent or iface.mainWindow()

    extent_rectangle = _troba_extent_topografic(iface)
    if extent_rectangle is None:
        QMessageBox.warning(
            win, "Sense referencial topogràfic carregat",
            "Cal carregar primer «Referencial topogràfic territorial "
            "vectorial» (Cartografia de referència), ja que les "
            "qualificacions especials es retallen a la mateixa àrea."
        )
        return

    ambit_lyr = _get_layer(LAYER_AMBIT_NAME)
    if ambit_lyr is None or ambit_lyr.featureCount() == 0:
        QMessageBox.warning(
            win, "Sense àmbit de l'IOF",
            "Crea primer l'àmbit de l'IOF (Cadastre → Crear àmbit de "
            "l'IOF) -- necessari per limitar FAUNA, UP i ZAU a la zona "
            "rellevant."
        )
        return
    extent_ambit = ambit_lyr.extent()

    proj_path = QgsProject.instance().absolutePath()
    if not proj_path:
        from .iof_utils import ensure_project_saved
        proj_path = ensure_project_saved(win)
        if not proj_path:
            return

    gpkg_path = os.path.join(proj_path, "IOF_Qualificacions.gpkg")

    # Rectangle sencer, només contorn
    FONT_ENPE = ("ENPE", "wfs", _wfs_uri(WFS_ESPAIS_NATURALS, "ESPAIS_NATURALS:ESPAISNATURALS_ENPE"), "0,167,126,255")
    FONT_PEIN = ("PEIN", "wfs", _wfs_uri(WFS_ESPAIS_NATURALS, "ESPAIS_NATURALS:ESPAISNATURALS_PEIN"))
    FONT_N2000 = ("Xarxa_Natura_2000", "wfs", _wfs_uri(WFS_ESPAIS_NATURALS, "ESPAIS_NATURALS:ESPAISNATURALS_XARNAT_2000"))
    # PPP es descarrega diferent (SHP), es tracta a part més avall

    # Només dins l'Àmbit IOF, amb contorn també (colors propis)
    FONTS_AMBIT = [
        ("UP", "wfs", _wfs_uri(WFS_ESPAIS_NATURALS, "ESPAIS_NATURALS:ESPAISNATURALS_FORESTS"), "136,14,79,255", "dot"),
        ("FAUNA", "wfs", _wfs_uri(WFS_FAUNA, "FAUNA:FAUNA_AIFF_PUBLICA"), "255,214,0,255", "dash"),
    ]

    QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
    n_afectacions = {}  # codi -> nombre d'afectacions (features trobades)
    errors = []
    es_primera = True
    try:
        proj = QgsProject.instance()
        grup = _grup_qualificacions(proj)
        grup_altres = _grup_altres_qualificacions(proj)

        # Capa de memòria que va acumulant tot el que ja s'ha mostrat amb
        # prioritat més alta -- jerarquia demanada per l'usuari: ENPE >
        # PEIN-PE > PEIN > LIC-ZEPA > ZEPA > LIC. Com que PEIN-PE i PEIN
        # ja comparteixen la mateixa simbologia (decisió anterior), es
        # tracten com un sol nivell de jerarquia (subtreure'ls per
        # separat donaria el mateix resultat visual, ja que
        # (A∪B)-C = (A-C)∪(B-C)). Només Xarxa Natura 2000 (LIC-ZEPA/
        # ZEPA/LIC, amb estils diferents) cal separar-la en 3 subcapes.
        cobert_lyr = QgsVectorLayer("Polygon?crs=EPSG:25831", "cobert_jerarquia", "memory")

        def _afegeix_a_cobert(layer_nou):
            cobert_lyr.startEditing()
            for feat in layer_nou.getFeatures():
                geom = feat.geometry()
                if geom and not geom.isEmpty():
                    nf = QgsFeature(cobert_lyr.fields())
                    nf.setGeometry(geom)
                    cobert_lyr.addFeature(nf)
            cobert_lyr.commitChanges()

        def _subtreu_cobert(layer):
            if cobert_lyr.featureCount() == 0:
                return layer
            try:
                import processing
                res = processing.run("native:difference", {
                    "INPUT": layer, "OVERLAY": cobert_lyr, "OUTPUT": "TEMPORARY_OUTPUT",
                })
                diferencia = res["OUTPUT"]
                # IMPORTANT: native:difference entre dues capes que
                # haurien de coincidir exactament (p. ex. la mateixa
                # zona protegida, però digitalitzada per separat a
                # ENPE/PEIN/Xarxa Natura 2000, amb vèrtexs lleugerament
                # diferents) sol deixar residus minúsculs ("slivers")
                # de només uns pocs m² -- una diferència de precisió
                # numèrica, no una àrea real visible. Es filtren fora
                # (llindar 100 m² = 0,01 ha) perquè no es mostrin amb
                # estil ni etiqueta com si fossin una zona real.
                res_filtrat = processing.run("native:extractbyexpression", {
                    "INPUT": diferencia, "EXPRESSION": "$area > 100",
                    "OUTPUT": "TEMPORARY_OUTPUT",
                })
                return res_filtrat["OUTPUT"]
            except Exception:
                return layer

        # --- ENPE: rectangle sencer, només contorn (color únic) -- prioritat 1 ---
        nom, provider, uri, color = FONT_ENPE
        try:
            lyr = QgsVectorLayer(uri, nom, provider.upper())
            if not lyr.isValid():
                errors.append(nom + " (no s'ha pogut connectar amb el servei WFS)")
            else:
                retallat = _retalla_per_extent(lyr, extent_rectangle)
                n_afectacions[nom] = retallat.featureCount()
                if retallat.featureCount() > 0:
                    _escriu_capa_gpkg(retallat, gpkg_path, nom, es_primera)
                    es_primera = False
                    lyr_carregada = QgsVectorLayer(gpkg_path + "|layername=" + nom, NOMS_LLEGENDA.get(nom, nom), "ogr")
                    if lyr_carregada.isValid():
                        camp_nom = _troba_camp_output(lyr_carregada, "NOM_ESPAI") or "NOM_ESPAI"
                        # Equivalent a "Parque Natural" del catàleg MTN25: discontinu-puntejat
                        _mostra_amb_vores_reals(
                            proj, grup, lyr_carregada, extent_rectangle, color, "dash dot",
                            camp_etiqueta=camp_nom if camp_nom in lyr_carregada.fields().names() else None,
                            color_etiqueta=color,
                        )
                        _aplica_metadata(lyr_carregada, nom)
                        _afegeix_a_cobert(lyr_carregada)
        except Exception as e:
            _log_error("ENPE", e)
            n_afectacions.pop(nom, None)
            errors.append(nom + " (error en carregar -- revisa la connexió a internet)")

        # --- PEIN-PE (prioritat 2) > PEIN (prioritat 3): mateixa simbologia, jerarquia separada ---
        nom, provider, uri = FONT_PEIN
        try:
            lyr = QgsVectorLayer(uri, nom, provider.upper())
            if not lyr.isValid():
                errors.append(nom + " (no s'ha pogut connectar amb el servei WFS)")
            else:
                retallat = _retalla_per_extent(lyr, extent_rectangle)
                n_pein = 0
                if retallat.featureCount() > 0:
                    _escriu_capa_gpkg(retallat, gpkg_path, nom, es_primera)
                    es_primera = False
                    lyr_gpkg = QgsVectorLayer(gpkg_path + "|layername=" + nom, NOMS_LLEGENDA.get(nom, nom), "ogr")
                    if lyr_gpkg.isValid():
                        planif_field = _troba_camp_output(lyr_gpkg, "PLANIF") or "PLANIF"
                        camp_nom = _troba_camp_output(lyr_gpkg, "NOM_PEIN") or "NOM_PEIN"
                        expr_sense_pla = f'"{planif_field}" IS NULL OR "{planif_field}" = \'Sense planificació\''

                        # Mateixa simbologia (equivalent a "Parque Nacional" del
                        # catàleg MTN25: discontinu) per a totes dues -- només
                        # canvia l'ordre de jerarquia (PEIN-PE abans que PEIN)
                        subtipus = [
                            ("PEIN-PE", "NOT (" + expr_sense_pla + ")"),
                            ("PEIN", expr_sense_pla),
                        ]
                        for subnom, expr in subtipus:
                            subset = _filtra_per_expressio(lyr_gpkg, expr)
                            if subset is None or subset.featureCount() == 0:
                                continue
                            n_pein += subset.featureCount()
                            visible = _subtreu_cobert(subset)
                            nom_llegenda = NOMS_LLEGENDA_SUBTIPUS.get(subnom, subnom)
                            if visible.featureCount() > 0:
                                # No queda per sota de cap altra qualificació
                                # ja mostrada: es veu una part pròpia i
                                # distintiva -- va al grup principal, visible.
                                capa_final, grup_final, es_visible = visible, grup, True
                            else:
                                # Queda completament per sota d'una altra
                                # (superfície ja coberta per una qualificació
                                # de prioritat més alta): la forma retallada
                                # queda buida, no es pot veure. Es mostra la
                                # forma ORIGINAL sencera (sense retallar per
                                # jerarquia) a "Altres qualificacions
                                # especials", oculta, per deixar constància
                                # que la finca també hi és afectada.
                                capa_final, grup_final, es_visible = subset, grup_altres, False
                            # Escriu la subcapa (resultat de processing,
                            # temporal per naturalesa) al GeoPackage i la
                            # recarrega des de disc -- mateix mètode que
                            # ENPE/PPP/FAUNA, perquè no quedi com a capa
                            # temporal (scratch) al projecte.
                            gpkg_nom = "vis_" + subnom.replace("-", "_")
                            _escriu_capa_gpkg(capa_final, gpkg_path, gpkg_nom, False)
                            capa_disc = QgsVectorLayer(
                                gpkg_path + "|layername=" + gpkg_nom, nom_llegenda, "ogr"
                            )
                            if capa_disc.isValid():
                                capa_final = capa_disc
                            else:
                                capa_final.setName(nom_llegenda)
                            _mostra_amb_vores_reals(
                                proj, grup_final, capa_final, extent_rectangle, "0,167,126,255", "dash",
                                camp_etiqueta=camp_nom if camp_nom in capa_final.fields().names() else None,
                                color_etiqueta="0,167,126,255",
                                visible=es_visible,
                            )
                            _afegeix_a_cobert(subset)
                n_afectacions[nom] = n_pein
        except Exception as e:
            _log_error("PEIN", e)
            n_afectacions.pop(nom, None)
            errors.append(nom + " (error en carregar -- revisa la connexió a internet)")

        # --- Xarxa Natura 2000: LIC-ZEPA (prioritat 4) > ZEPA (prioritat 5) > LIC (prioritat 6) ---
        nom, provider, uri = FONT_N2000
        try:
            lyr = QgsVectorLayer(uri, nom, provider.upper())
            if not lyr.isValid():
                errors.append(nom + " (no s'ha pogut connectar amb el servei WFS)")
            else:
                retallat = _retalla_per_extent(lyr, extent_rectangle)
                n_n2000 = 0
                if retallat.featureCount() > 0:
                    _escriu_capa_gpkg(retallat, gpkg_path, nom, es_primera)
                    es_primera = False
                    lyr_gpkg = QgsVectorLayer(gpkg_path + "|layername=" + nom, NOMS_LLEGENDA.get(nom, nom), "ogr")
                    if lyr_gpkg.isValid():
                        camp_lic = _troba_camp_output(lyr_gpkg, "LIC_ZEC") or "LIC_ZEC"
                        camp_zepa = _troba_camp_output(lyr_gpkg, "ZEPA") or "ZEPA"
                        camp_nom = _troba_camp_output(lyr_gpkg, "NOM_XN2") or "NOM_XN2"

                        subtipus = [
                            ("LIC-ZEPA", f'"{camp_lic}" = \'Sí\' AND "{camp_zepa}" = \'Sí\'', "20,80,140,255", "dash dot dot"),
                            ("ZEPA", f'"{camp_lic}" <> \'Sí\' AND "{camp_zepa}" = \'Sí\'', "100,170,220,255", "dot"),
                            ("LIC", f'"{camp_lic}" = \'Sí\' AND "{camp_zepa}" <> \'Sí\'', "31,120,180,255", "dash"),
                        ]
                        for subnom, expr, subcolor, substil in subtipus:
                            subset = _filtra_per_expressio(lyr_gpkg, expr)
                            if subset is None or subset.featureCount() == 0:
                                continue
                            n_n2000 += subset.featureCount()
                            visible = _subtreu_cobert(subset)
                            nom_llegenda = NOMS_LLEGENDA_SUBTIPUS.get(subnom, subnom)
                            if visible.featureCount() > 0:
                                capa_final, grup_final, es_visible = visible, grup, True
                            else:
                                capa_final, grup_final, es_visible = subset, grup_altres, False
                            gpkg_nom = "vis_" + subnom.replace("-", "_")
                            _escriu_capa_gpkg(capa_final, gpkg_path, gpkg_nom, False)
                            capa_disc = QgsVectorLayer(
                                gpkg_path + "|layername=" + gpkg_nom, nom_llegenda, "ogr"
                            )
                            if capa_disc.isValid():
                                capa_final = capa_disc
                            else:
                                capa_final.setName(nom_llegenda)
                            _mostra_amb_vores_reals(
                                proj, grup_final, capa_final, extent_rectangle, subcolor, substil,
                                camp_etiqueta=camp_nom if camp_nom in capa_final.fields().names() else None,
                                color_etiqueta=subcolor,
                                visible=es_visible,
                            )
                            _afegeix_a_cobert(subset)
                n_afectacions[nom] = n_n2000
        except Exception as e:
            _log_error("Xarxa Natura 2000", e)
            n_afectacions.pop(nom, None)
            errors.append(nom + " (error en carregar -- revisa la connexió a internet)")

        # --- PPP: descàrrega SHP, rectangle sencer, contorn amb etiqueta ---
        try:
            shp_path = _descarrega_shp_zip(URL_PPP, "ppp")
            lyr = QgsVectorLayer(shp_path, "PPP", "ogr")
            if not lyr.isValid():
                errors.append("PPP (no s'ha pogut connectar amb el servei)")
            else:
                retallat = _retalla_per_extent(lyr, extent_rectangle)
                n_afectacions["PPP"] = retallat.featureCount()
                if retallat.featureCount() > 0:
                    _escriu_capa_gpkg(retallat, gpkg_path, "PPP", es_primera)
                    es_primera = False
                    lyr_carregada = QgsVectorLayer(
                        gpkg_path + "|layername=PPP", NOMS_LLEGENDA["PPP"], "ogr"
                    )
                    if lyr_carregada.isValid():
                        camp_nom = _troba_camp_output(lyr_carregada, "NOM") or "NOM"
                        _mostra_amb_vores_reals(
                            proj, grup_altres, lyr_carregada, extent_rectangle, "204,85,0,255", "dash dot",
                            gruix_linia=0.8,
                            camp_etiqueta=camp_nom if camp_nom in lyr_carregada.fields().names() else None,
                            color_etiqueta="204,85,0,255",
                            visible=False,
                        )
                        _aplica_metadata(lyr_carregada, "PPP")
        except Exception as e:
            _log_error("PPP", e)
            n_afectacions.pop("PPP", None)
            errors.append("PPP (error en carregar -- revisa la connexió a internet)")

        # --- ZAU: descàrrega SHP, es mostra si toca l'Àmbit IOF, retallat al rectangle sencer ---
        try:
            shp_path = _descarrega_shp_zip(URL_ZAU, "zau")
            lyr = QgsVectorLayer(shp_path, "ZAU", "ogr")
            if not lyr.isValid():
                errors.append("ZAU (no s'ha pogut connectar amb el servei)")
            else:
                candidats = _retalla_per_extent(lyr, extent_rectangle, clip=False)
                inclosos = _selecciona_per_interseccio_real(candidats, ambit_lyr)
                n_afectacions["ZAU"] = inclosos.featureCount()
                if inclosos.featureCount() > 0:
                    retallat = _retalla_per_extent(inclosos, extent_rectangle)
                    _escriu_capa_gpkg(retallat, gpkg_path, "ZAU", es_primera)
                    es_primera = False
                    lyr_carregada = QgsVectorLayer(
                        gpkg_path + "|layername=ZAU", NOMS_LLEGENDA["ZAU"], "ogr"
                    )
                    if lyr_carregada.isValid():
                        _mostra_amb_vores_reals(
                            proj, grup_altres, lyr_carregada, extent_rectangle, "66,66,66,255", "dash dot",
                            visible=False,
                        )
                        _aplica_metadata(lyr_carregada, "ZAU")
        except Exception as e:
            _log_error("ZAU", e)
            n_afectacions.pop("ZAU", None)
            errors.append("ZAU (error en carregar -- revisa la connexió a internet)")

        # --- UP, FAUNA: es mostren si toquen l'Àmbit IOF, retallats al rectangle sencer ---
        # FAUNA va al grup principal (visible); UP sempre a "Altres" (ocult).
        for nom, provider, uri, color, estil in FONTS_AMBIT:
            try:
                lyr = QgsVectorLayer(uri, nom, provider.upper())
                if not lyr.isValid():
                    errors.append(nom + " (no s'ha pogut connectar amb el servei WFS)")
                    continue
                candidats = _retalla_per_extent(lyr, extent_rectangle, clip=False)
                inclosos = _selecciona_per_interseccio_real(candidats, ambit_lyr)
                retallat = _retalla_per_extent(inclosos, extent_rectangle)
                if nom == "FAUNA":
                    # No es mostren ni es calculen espècies sense
                    # categoria de protecció assignada (camp PROT_CAT
                    # buit) -- si no hi ha protecció informada, no
                    # representa una qualificació especial real.
                    # S'aplica DESPRÉS del retall (no amb
                    # setSubsetString sobre el WFS directament: aquest
                    # mètode retorna 0 elements amb aquest servidor --
                    # provat en directe --, probablement perquè el
                    # filtre no es tradueix correctament a una petició
                    # OGC vàlida pel proveïdor WFS).
                    camp_prot = _troba_camp_output(retallat, "PROT_CAT") or "PROT_CAT"
                    if camp_prot in retallat.fields().names():
                        import processing
                        res_filtrat = processing.run("native:extractbyexpression", {
                            "INPUT": retallat,
                            "EXPRESSION": '"' + camp_prot + '" IS NOT NULL AND "' + camp_prot + '" <> \'\'',
                            "OUTPUT": "TEMPORARY_OUTPUT",
                        })
                        retallat = res_filtrat["OUTPUT"]
                n_afectacions[nom] = retallat.featureCount()
                if retallat.featureCount() == 0:
                    continue
                _escriu_capa_gpkg(retallat, gpkg_path, nom, es_primera)
                es_primera = False
                lyr_carregada = QgsVectorLayer(
                    gpkg_path + "|layername=" + nom, NOMS_LLEGENDA.get(nom, nom), "ogr"
                )
                if lyr_carregada.isValid():
                    camp_etiqueta = None
                    es_expr = False
                    color_halo_et = "255,255,255,255"
                    if nom == "FAUNA":
                        camp_esp = _troba_camp_output(lyr_carregada, "NOM_ESP") or "NOM_ESP"
                        camp_prot = _troba_camp_output(lyr_carregada, "PROT_CAT") or "PROT_CAT"
                        noms_capa = lyr_carregada.fields().names()
                        # Halo gris fosc (en lloc del blanc per defecte):
                        # el groc de FAUNA té molt poc contrast amb un halo
                        # blanc i costa de llegir.
                        color_halo_et = "80,80,80,255"
                        if camp_esp in noms_capa and camp_prot in noms_capa:
                            camp_etiqueta = (
                                '"' + camp_esp + '"' +
                                " || CASE WHEN \"" + camp_prot + "\" IS NOT NULL AND " +
                                "\"" + camp_prot + "\" <> '' THEN ' (' || \"" + camp_prot +
                                "\" || ')' ELSE '' END"
                            )
                            es_expr = True
                        elif camp_esp in noms_capa:
                            camp_etiqueta = camp_esp
                    grup_final = grup if nom == "FAUNA" else grup_altres
                    _mostra_amb_vores_reals(
                        proj, grup_final, lyr_carregada, extent_rectangle, color, estil,
                        camp_etiqueta=camp_etiqueta, color_etiqueta=color,
                        color_halo_etiqueta=color_halo_et, es_expressio_etiqueta=es_expr,
                        visible=(nom == "FAUNA"),
                    )
                    _aplica_metadata(lyr_carregada, nom)
            except Exception as e:
                _log_error(nom, e)
                n_afectacions.pop(nom, None)
                errors.append(nom + " (error en carregar -- revisa la connexió a internet)")

        # LU: NO es visualitza al mapa -- el càlcul real es fa per punt
        # amb GetFeatureInfo a "Exportar qualificacions especials".

        iface.mapCanvas().setExtent(extent_rectangle)
        iface.mapCanvas().refresh()
    finally:
        QApplication.restoreOverrideCursor()

    # Noms de visualització i ordre per al resum -- diferent de
    # NOMS_LLEGENDA (que fa servir el nom sencer per a la llegenda del
    # mapa): aquí es vol el codi curt (o "fauna" en minúscula, ja que
    # no és pròpiament una sigla), per fer bullets compactes agrupant
    # totes les qualificacions amb el mateix nombre d'afectacions.
    NOMS_DISPLAY_RESUM = {
        "ENPE": "ENPE", "PEIN": "PEIN", "Xarxa_Natura_2000": "Xarxa Natura 2000",
        "UP": "UP", "FAUNA": "fauna", "ZAU": "ZAU",
    }
    ORDRE_CODIS_RESUM = ["ENPE", "PEIN", "Xarxa_Natura_2000", "ZAU", "UP", "FAUNA"]

    def _uneix_amb_i(items):
        items = list(items)
        if len(items) <= 1:
            return items[0] if items else ""
        return ", ".join(items[:-1]) + " i " + items[-1]

    def _text_afectacions(n):
        return "1 afectació" if n == 1 else str(n) + " afectacions"

    grups_per_n = {}
    for codi in ORDRE_CODIS_RESUM:
        if codi in n_afectacions:
            grups_per_n.setdefault(n_afectacions[codi], []).append(NOMS_DISPLAY_RESUM[codi])

    bullets = []
    for n in sorted(grups_per_n, reverse=True):
        bullets.append(
            _uneix_amb_i(grups_per_n[n]) + " (" + _text_afectacions(n) + " de qualificació especial)"
        )
    if "PPP" in n_afectacions:
        bullets.append("PPP (" + _text_afectacions(n_afectacions["PPP"]) + " de PPP)")

    text = "Capes carregades correctament (" + str(len(n_afectacions)) + "):\n"
    text += "\n".join("  • " + b for b in bullets)

    text += "\n\nCapes no carregades:\n"
    text += "  • BS (cartografia no disponible)"

    if errors:
        text += "\n\nNo s'han pogut carregar (" + str(len(errors)) + "):\n"
        text += "\n".join("  • " + n for n in errors)

    text += ("\n\nNota: LU no es visualitza al mapa (el càlcul es fa per "
             "punt a \"Exportar qualificacions especials\").")

    text += "\n\nGeoPackage desat a:\n" + gpkg_path
    QMessageBox.information(win, "Qualificacions especials", text)


# ---------------------------------------------------------------------------
# Botó 2: "Exportar qualificacions especials" -- calcular + generar informe
# ---------------------------------------------------------------------------

def _clip_i_interseca(espais_layer, ambit_layer, unitats_layer):
    """Retalla espais_layer a l'àmbit i el creua amb les tipologies
    forestals. Retorna la capa resultant de la intersecció (en memòria)."""
    import processing
    res_clip = processing.run("native:clip", {
        "INPUT": espais_layer, "OVERLAY": ambit_layer,
        "OUTPUT": "TEMPORARY_OUTPUT",
    })
    clipped = res_clip["OUTPUT"]
    if clipped.featureCount() == 0:
        return None
    res_int = processing.run("native:intersection", {
        "INPUT": unitats_layer, "OVERLAY": clipped,
        "OUTPUT": "TEMPORARY_OUTPUT",
    })
    return res_int["OUTPUT"]


def _troba_camp_output(output_layer, field_name):
    if not field_name:
        return None
    names = output_layer.fields().names()
    if field_name in names:
        return field_name
    for n in names:
        if n.startswith(field_name + "_") or n.startswith(field_name + "2"):
            return n
    return None


def _acumula_interseccio(interseccio, codi_field, qualif_fn, resum_total,
                          resum_per_unitat):
    """Recorre les entitats d'una intersecció i acumula superfície (ha) a
    resum_total (per qualificació) i resum_per_unitat (per unitat+qualif).
    qualif_fn(feature) -> nom de qualificació a assignar."""
    n = 0
    for feat in interseccio.getFeatures():
        geom = feat.geometry()
        if not geom or geom.isEmpty():
            continue
        area_ha = geom.area() / 10000.0
        if area_ha <= 0:
            continue
        codi = str(feat[codi_field]) if codi_field else ""
        qualif = qualif_fn(feat)
        if not qualif:
            continue
        resum_total[qualif] = resum_total.get(qualif, 0.0) + area_ha
        key = (codi, qualif)
        resum_per_unitat[key] = resum_per_unitat.get(key, 0.0) + area_ha
        n += 1
    return n


def _acumula_interseccio_amb_nom(interseccio, codi_field, qualif_fn, nom_fn,
                                  resum_total, resum_per_unitat, noms_trobats,
                                  geometries_per_unitat=None):
    """
    Com _acumula_interseccio, però acumula per (qualificació, nom concret
    de l'espai) en lloc de només per qualificació -- necessari per poder
    cercar la correspondència de cada espai concret amb el desplegable
    del formulari PDF.
    - resum_total: (qualificacio, nom_espai) -> area_ha
    - resum_per_unitat: (codi_unitat, qualificacio, nom_espai) -> area_ha
    - noms_trobats: set de (qualificacio, nom_espai) trobats (per fer
      la cerca de correspondència del PDF una sola vegada per cadascun)
    - geometries_per_unitat (opcional): codi_unitat -> [QgsGeometry, ...],
      acumula la geometria real de cada tall (no només l'àrea) perquè
      després es pugui calcular la unió real per unitat -- necessari
      per no comptar dues vegades la superfície on coincideixen
      diverses qualificacions (p. ex. una zona que és alhora ENPE i
      PEIN no pot sumar el doble de superfície afectada).
    qualif_fn(feature) -> codi de qualificació ("ENPE", "PEIN", etc.)
    nom_fn(feature) -> nom concret de l'espai (p. ex. NOM_ESPAI)
    """
    n = 0
    for feat in interseccio.getFeatures():
        geom = feat.geometry()
        if not geom or geom.isEmpty():
            continue
        area_ha = geom.area() / 10000.0
        if area_ha <= 0:
            continue
        codi = str(feat[codi_field]) if codi_field else ""
        qualif = qualif_fn(feat)
        if not qualif:
            continue
        nom_espai = nom_fn(feat) if nom_fn else None
        nom_espai = (nom_espai or "").strip() or "(sense nom)"
        key_total = (qualif, nom_espai)
        resum_total[key_total] = resum_total.get(key_total, 0.0) + area_ha
        key_unitat = (codi, qualif, nom_espai)
        resum_per_unitat[key_unitat] = resum_per_unitat.get(key_unitat, 0.0) + area_ha
        noms_trobats.add(key_total)
        if geometries_per_unitat is not None:
            geometries_per_unitat.setdefault(codi, []).append(QgsGeometry(geom))
        n += 1
    return n


def _calcula_superficie_unio(geometries):
    """Retorna la superfície (ha) de la unió real d'una llista de
    geometries -- evita comptar dues vegades la part on se superposen."""
    if not geometries:
        return 0.0
    try:
        unio = QgsGeometry.unaryUnion(geometries)
        return unio.area() / 10000.0 if unio and not unio.isEmpty() else 0.0
    except Exception:
        # Si unaryUnion falla (geometries invalides), combinem de una en una
        unio = None
        for g in geometries:
            unio = g if unio is None else unio.combine(g)
        return unio.area() / 10000.0 if unio and not unio.isEmpty() else 0.0


def _get_feature_info_muc(x, y, crs="EPSG:25831", d=1):
    """Consulta puntual GetFeatureInfo al WMS del Mapa Urbanístic de
    Catalunya. Retorna una tupla (codi_qual_muc, desc_qual_ajunt), o
    (None, None) si el punt no té cap qualificació urbanística
    trobada. `desc_qual_ajunt` és la descripció local/municipal
    concreta (p. ex. "Àrea de Ribera amb Regulació Específica") -- NO
    correspon necessàriament a un dels 12 noms de parc del desplegable
    del PDF (que són parcs comarcals/metropolitans, gestionats per ens
    supramunicipals), però és la identificació més concreta disponible
    en aquest servei per citar a Observacions.

    NO es captura cap excepció aquí (abans es feia, retornant
    (None, None) tant si no hi havia cap dada al punt com si el servei
    no responia): un error de xarxa/servei es propaga perquè qui ho
    crida ho pugui distingir d'un punt genuinament sense afectació --
    altrament un servei caigut es confonia amb "cap unitat afectada
    per LU", donant un resultat fals per bo sense cap avís."""
    bbox = str(x - d) + "," + str(y - d) + "," + str(x + d) + "," + str(y + d)
    params = {
        "SERVICE": "WMS", "VERSION": "1.3.0", "REQUEST": "GetFeatureInfo",
        "LAYERS": "MUC_4QUAL", "QUERY_LAYERS": "MUC_4QUAL", "STYLES": "",
        "CRS": crs, "BBOX": bbox, "WIDTH": "3", "HEIGHT": "3",
        "I": "1", "J": "1", "INFO_FORMAT": "text/xml",
    }
    url = WMS_MUC + "?" + urllib.parse.urlencode(params)
    if not url.lower().startswith("https://"):
        raise ValueError("URL no permesa (només HTTPS): " + url)
    req = urllib.request.Request(url, headers=USER_AGENT)
    # URL validada com a HTTPS just abans; mai ve de l'usuari (constant fixa)
    with urllib.request.urlopen(req, timeout=15) as resp:  # nosec B310
        data = resp.read().decode("utf-8", errors="replace")
    m_codi = re.search(r'Name="CODI_QUAL_MUC">([^<]+)<', data)
    m_desc = re.search(r'Name="DESC_QUAL_AJUNT">([^<]+)<', data)
    codi = m_codi.group(1) if m_codi else None
    desc = m_desc.group(1) if m_desc else None
    return codi, desc


def _troba_risc_incendi_tipus(ambit_lyr):
    """Descarrega el mapa de risc d'incendi tipus de Catalunya (Centre de
    la Propietat Forestal), el retalla a l'àmbit de l'IOF, i retorna el
    conjunt de valors de RISC trobats (normalment només un, ja que les
    zones són molt grans -- 90 per a tot Catalunya).

    NO es captura cap excepció aquí (abans es feia, retornant un set
    buit en qualsevol error): si el servei no respon o el fitxer
    descarregat no és vàlid, es propaga l'excepció perquè qui la crida
    ho pugui detectar i avisar-ne -- altrament un servei caigut es
    confonia amb "cap zona de risc trobada", donant per bo un resultat
    que en realitat no s'havia pogut calcular."""
    shp_path = _descarrega_shp_zip(URL_RISC_INCENDI_TIPUS, "riscincendi")
    lyr = QgsVectorLayer(shp_path, "risc_incendi_tipus", "ogr")
    if not lyr.isValid():
        raise RuntimeError("El fitxer de risc d'incendi tipus descarregat no és una capa vàlida")
    import processing
    res = processing.run("native:clip", {
        "INPUT": lyr, "OVERLAY": ambit_lyr, "OUTPUT": "TEMPORARY_OUTPUT",
    })
    clipped = res["OUTPUT"]
    valors = set()
    for feat in clipped.getFeatures():
        geom = feat.geometry()
        if geom and not geom.isEmpty() and feat["RISC"]:
            valors.add(str(feat["RISC"]).strip())
    return valors


def _confirma_abans_exportar(win, consultades_ok, errors):
    """
    Mostra un resum de què s'ha pogut calcular i què no abans de generar
    l'informe (mateix patró que la revisió abans d'exportar el TXT de
    l'IOF): ✅ correctes, ⚠️ no consultades (error de connexió), i
    🔲 sense cartografia disponible (BS, sempre exclosa). Demana
    confirmació abans de continuar.
    """
    text = "Abans de generar l'informe, revisa què s'ha pogut calcular:\n\n"
    if consultades_ok:
        text += "✅ Calculades correctament:\n"
        for n in consultades_ok:
            text += "  • " + n + "\n"
        text += "\n"
    if errors:
        text += "⚠️ No s'han pogut consultar (revisa la connexió a internet):\n"
        for n in errors:
            text += "  • " + n + "\n"
        text += "\n"
    text += "🔲 Sense cartografia disponible:\n"
    text += "  • Boscos Singulars\n\n"
    text += (
        "⚠️ Important: la superfície afectada per unitat d'actuació o "
        "rodal només serà correcta si totes les tipologies forestals "
        "del mapa estan digitalitzades. Si en falta alguna, el "
        "resultat per unitat no reflectirà la superfície real "
        "afectada.\n\n"
    )
    text += "Vols continuar i generar l'informe?"

    resposta = QMessageBox.question(
        win, "Revisió abans d'exportar", text,
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        QMessageBox.StandardButton.No,
    )
    return resposta == QMessageBox.StandardButton.Yes


def _troba_gpkg_qualificacions():
    """Retorna el path de IOF_Qualificacions.gpkg del projecte actual
    (generat per "Qualificacions especials afectades"), o None si
    encara no existeix.

    Es busca primer a l'arrel de la carpeta del projecte (ubicació
    actual); si no hi és, es comprova també la subcarpeta
    "qualificacions" (ubicació antiga, per compatibilitat amb
    projectes generats abans d'aquest canvi)."""
    proj_path = QgsProject.instance().absolutePath()
    if not proj_path:
        return None
    gpkg_arrel = os.path.join(proj_path, "IOF_Qualificacions.gpkg")
    if os.path.exists(gpkg_arrel):
        return gpkg_arrel
    gpkg_antic = os.path.join(proj_path, "qualificacions", "IOF_Qualificacions.gpkg")
    return gpkg_antic if os.path.exists(gpkg_antic) else None


def _carrega_capa_font(nom_qualif, gpkg_path, wfs_url, wfs_typename):
    """
    Retorna (layer, ja_retallada) per a una qualificació. Prioritza la
    capa ja retallada dins IOF_Qualificacions.gpkg (generada per
    "Qualificacions especials afectades" -- molt més ràpid, ja que
    només conté l'àrea rellevant); si no hi és (p. ex. no s'hi ha
    trobat cap element en aquella zona, o encara no s'ha generat el
    GeoPackage), fa una descàrrega WFS completa com a alternativa
    (`ja_retallada=False`, caldrà retallar-la després amb l'àmbit).
    """
    if gpkg_path:
        uri = gpkg_path + "|layername=" + nom_qualif
        lyr = QgsVectorLayer(uri, nom_qualif, "ogr")
        if lyr.isValid():
            return lyr, True
    uri = _wfs_uri(wfs_url, wfs_typename)
    lyr = QgsVectorLayer(uri, nom_qualif + "_tmp", "WFS")
    return (lyr if lyr.isValid() else None), False


def _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr):
    """Intersecció amb les tipologies forestals. Si la capa ja ve
    retallada des del GeoPackage de qualificacions, es pot intersecar
    directament (ja és petita); si no, cal retallar-la primer a
    l'àmbit de l'IOF per rendiment."""
    if not ja_retallada:
        return _clip_i_interseca(lyr, ambit_lyr, unitats_lyr)
    if lyr.featureCount() == 0:
        return None
    import processing
    res = processing.run("native:intersection", {
        "INPUT": unitats_lyr, "OVERLAY": lyr, "OUTPUT": "TEMPORARY_OUTPUT",
    })
    return res["OUTPUT"]



def run_exportar_qualificacions(iface, parent=None):
    """
    Calcula la superfície de cada tipologia forestal afectada per
    cadascuna de les qualificacions especials (per nom concret d'espai),
    indica si cada espai té correspondència amb el desplegable del
    formulari oficial del PTGMF (i, si no, el nom real per citar a
    Observacions), i genera un informe (Excel o text) amb el resultat
    per unitat i el total per a la finca.

    Particularitats:
    - FAUNA: només s'indica si la finca està afectada o no, més la
      llista d'espècies concretes trobades (per citar a Observacions,
      ja que el PDF només té l'opció genèrica "FAUNA").
    - UP: només es tenen en compte les forests amb número de Catàleg
      (CUP) informat -- es reporta com "Bosc protector" + nom + CUP.
    - PPP: NO és una qualificació especial (és informació de prevenció
      d'incendis) -- s'exclou del càlcul per tipologia forestal, però
      sí es calcula el total per a la finca sencera.
    - LU: es manté el mecanisme per punt (GetFeatureInfo) -- només
      indica si cada unitat hi és afectada o no, no el nom del parc
      concret (limitació coneguda, sense resoldre).
    - BS: sense cartografia, no s'inclou en cap càlcul.
    """
    win = parent or iface.mainWindow()

    unitats_lyr = _troba_unitats_layer()
    if unitats_lyr is None or unitats_lyr.featureCount() == 0:
        QMessageBox.warning(
            win, "Sense tipologies forestals",
            "Digitalitza primer les tipologies forestals."
        )
        return
    codi_field = _troba_camp_codi(unitats_lyr)
    if not codi_field:
        QMessageBox.warning(
            win, "Camp no trobat",
            "La capa de tipologies forestals no té cap camp «codi_ua» ni "
            "«codi_rodal»."
        )
        return

    ambit_lyr = _get_layer(LAYER_AMBIT_NAME)
    if ambit_lyr is None or ambit_lyr.featureCount() == 0:
        QMessageBox.warning(
            win, "Sense àmbit de l'IOF",
            "Crea primer l'àmbit de l'IOF (Cadastre → Crear àmbit de l'IOF)."
        )
        return

    resum_total = {}      # (qualif, nom_espai) -> area_ha
    resum_per_unitat = {} # (codi_unitat, qualif, nom_espai) -> area_ha
    noms_trobats = set()  # (qualif, nom_espai) -- per cercar correspondencia PDF
    geometries_per_unitat = {}  # codi_unitat -> [QgsGeometry, ...] (per calcular la unio real)
    errors = []
    consultades_ok = []
    fauna_especies = set()
    ppp_total_finca = 0.0
    ppp_per_nom = {}
    gpkg_path = _troba_gpkg_qualificacions()

    QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
    try:
        # --- 1. ENPE ---
        try:
            lyr, ja_retallada = _carrega_capa_font(
                "ENPE", gpkg_path, WFS_ESPAIS_NATURALS,
                "ESPAIS_NATURALS:ESPAISNATURALS_ENPE",
            )
            if lyr:
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter:
                    camp_nom = _troba_camp_output(inter, "NOM_ESPAI") or "NOM_ESPAI"
                    _acumula_interseccio_amb_nom(
                        inter, codi_field, lambda f: "ENPE",
                        lambda f, c=camp_nom: f[c] if c in f.fields().names() else None,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )
                consultades_ok.append("ENPE")
            else:
                errors.append("ENPE")
        except Exception:
            errors.append("ENPE")

        # --- 2. PEIN / PEIN-PE (distingit pel camp PLANIF) + RF ---
        # RF (04) NO ve d'ENPE -- es creua directament amb la capa PEIN,
        # identificat pels 3 noms concrets ja validats manualment
        # (Ribera de l'Ebre a Flix, Illa de Canet, Basses de l'Albera).
        try:
            lyr, ja_retallada = _carrega_capa_font(
                "PEIN", gpkg_path, WFS_ESPAIS_NATURALS,
                "ESPAIS_NATURALS:ESPAISNATURALS_PEIN",
            )
            if lyr:
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter:
                    planif_field = _troba_camp_output(inter, "PLANIF")
                    camp_nom = _troba_camp_output(inter, "NOM_PEIN") or "NOM_PEIN"

                    def qualif_pein(f, planif_field=planif_field):
                        if not planif_field:
                            return "PEIN"
                        v = f[planif_field]
                        if not v or str(v).strip() == "Sense planificació":
                            return "PEIN"
                        return "PEIN-PE"

                    _acumula_interseccio_amb_nom(
                        inter, codi_field, qualif_pein,
                        lambda f, c=camp_nom: f[c] if c in f.fields().names() else None,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )

                    def qualif_rf(f, c=camp_nom):
                        nom = f[c] if c in f.fields().names() else None
                        if nom and nom.strip().upper() in _NOMS_PEIN_QUE_SON_RF:
                            return "RF"
                        return None

                    _acumula_interseccio_amb_nom(
                        inter, codi_field, qualif_rf,
                        lambda f, c=camp_nom: f[c] if c in f.fields().names() else None,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )
                consultades_ok.append("PEIN i PEIN-PE")
            else:
                errors.append("PEIN/PEIN-PE")
        except Exception:
            errors.append("PEIN/PEIN-PE")

        # --- 3. Xarxa Natura 2000: LIC / ZEPA / LIC-ZEPA ---
        try:
            lyr, ja_retallada = _carrega_capa_font(
                "Xarxa_Natura_2000", gpkg_path, WFS_ESPAIS_NATURALS,
                "ESPAIS_NATURALS:ESPAISNATURALS_XARNAT_2000",
            )
            if lyr:
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter:
                    lic_field = _troba_camp_output(inter, "LIC_ZEC")
                    zepa_field = _troba_camp_output(inter, "ZEPA")
                    camp_nom = _troba_camp_output(inter, "NOM_XN2") or "NOM_XN2"

                    def qualif_n2000(f, lic_field=lic_field, zepa_field=zepa_field):
                        es_lic = lic_field and str(f[lic_field]).strip().lower() == "sí"
                        es_zepa = zepa_field and str(f[zepa_field]).strip().lower() == "sí"
                        if es_lic and es_zepa:
                            return "LIC-ZEPA"
                        if es_lic:
                            return "LIC"
                        if es_zepa:
                            return "ZEPA"
                        return None

                    _acumula_interseccio_amb_nom(
                        inter, codi_field, qualif_n2000,
                        lambda f, c=camp_nom: f[c] if c in f.fields().names() else None,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )
                consultades_ok.append("Xarxa Natura 2000 (LIC / ZEPA / LIC-ZEPA)")
            else:
                errors.append("Xarxa Natura 2000")
        except Exception:
            errors.append("Xarxa Natura 2000")

        # --- 4. UP: només amb número de Catàleg (CUP) informat ---
        try:
            lyr, ja_retallada = _carrega_capa_font(
                "UP", gpkg_path, WFS_ESPAIS_NATURALS,
                "ESPAIS_NATURALS:ESPAISNATURALS_FORESTS",
            )
            if lyr:
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter:
                    camp_cup = _troba_camp_output(inter, "CUP")
                    camp_forest = _troba_camp_output(inter, "FOREST")
                    if camp_cup:
                        # El camp CUP és de tipus text i pot contenir codis com
                        # "202-A", "21-A", "92b", "81-bis" (números de catàleg
                        # vàlids, amb sufix de subparcel·la) o bé el text "No"
                        # (sense número de catàleg -- s'ha de considerar buit).
                        # Es considera vàlid quan comença per un dígit.
                        res_filtrat = _filtra_per_expressio(
                            inter,
                            '"' + camp_cup + '" IS NOT NULL AND '
                            'regexp_match("' + camp_cup + '", \'^[0-9]\') > 0'
                        )
                        if res_filtrat is not None:
                            inter = res_filtrat

                    def nom_up(f, camp_forest=camp_forest, camp_cup=camp_cup):
                        forest = f[camp_forest] if camp_forest and camp_forest in f.fields().names() else ""
                        cup = f[camp_cup] if camp_cup and camp_cup in f.fields().names() else ""
                        return "Bosc protector -- " + str(forest) + " (CUP " + str(cup) + ")"

                    _acumula_interseccio_amb_nom(
                        inter, codi_field, lambda f: "UP", nom_up,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )
                consultades_ok.append("UP")
            else:
                errors.append("UP")
        except Exception:
            errors.append("UP")

        # --- 5. FAUNA: només si/no + espècies afectades ---
        try:
            lyr, ja_retallada = _carrega_capa_font(
                "FAUNA", gpkg_path, WFS_FAUNA, "FAUNA:FAUNA_AIFF_PUBLICA",
            )
            if lyr:
                # IMPORTANT: cal retallar espacialment (clip/intersecció)
                # ABANS d'aplicar el filtre per expressió sobre PROT_CAT,
                # no després. La capa WFS de FAUNA té
                # `restrictToRequestBBOX='1'` -- si es filtra per
                # expressió PRIMER (sense cap context espacial), el
                # proveïdor no limita correctament la petició pel bbox
                # rellevant i el resultat pot sortir buit o incomplet.
                # Verificat en directe (qgis-mcp): filtrant abans de
                # retallar, 0 elements trobats en una zona amb 11 de
                # debò; retallant primer i filtrant després, els 11 es
                # troben correctament.
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter and not ja_retallada:
                    camp_prot = _troba_camp_output(inter, "PROT_CAT") or "PROT_CAT"
                    if camp_prot in inter.fields().names():
                        res_filtrat = _filtra_per_expressio(
                            inter, '"' + camp_prot + '" IS NOT NULL AND "' + camp_prot + '" <> \'\''
                        )
                        if res_filtrat is not None:
                            inter = res_filtrat
                if inter:
                    camp_esp = _troba_camp_output(inter, "NOM_ESP")
                    for feat in inter.getFeatures():
                        geom = feat.geometry()
                        if not geom or geom.isEmpty():
                            continue
                        area_ha = geom.area() / 10000.0
                        if area_ha <= 0:
                            continue
                        codi = str(feat[codi_field]) if codi_field else ""
                        resum_total[("FAUNA", "")] = resum_total.get(("FAUNA", ""), 0.0) + area_ha
                        key = (codi, "FAUNA", "")
                        resum_per_unitat[key] = resum_per_unitat.get(key, 0.0) + area_ha
                        geometries_per_unitat.setdefault(codi, []).append(QgsGeometry(geom))
                        if camp_esp and camp_esp in feat.fields().names() and feat[camp_esp]:
                            fauna_especies.add(str(feat[camp_esp]).strip())
                consultades_ok.append("Fauna protegida")
            else:
                errors.append("FAUNA")
        except Exception:
            errors.append("FAUNA")

        # --- 6. ZAU (per unitat, com la resta) ---
        try:
            lyr = None
            ja_retallada = False
            if gpkg_path:
                uri = gpkg_path + "|layername=ZAU"
                cand = QgsVectorLayer(uri, "ZAU", "ogr")
                if cand.isValid():
                    lyr, ja_retallada = cand, True
            if lyr is None:
                shp_path = _descarrega_shp_zip(URL_ZAU, "zau")
                lyr = QgsVectorLayer(shp_path, "zau_tmp", "ogr")
            if lyr.isValid():
                inter = _obte_interseccio(lyr, ja_retallada, ambit_lyr, unitats_lyr)
                if inter:
                    camp_codi_zau = _troba_camp_output(inter, "CODI_ZAU") or "CODI_ZAU"
                    _acumula_interseccio_amb_nom(
                        inter, codi_field, lambda f: "ZAU",
                        lambda f, c=camp_codi_zau: f[c] if c in f.fields().names() else None,
                        resum_total, resum_per_unitat, noms_trobats,
                        geometries_per_unitat=geometries_per_unitat,
                    )
                consultades_ok.append("ZAU")
            else:
                errors.append("ZAU")
        except Exception:
            errors.append("ZAU")

        # --- 7. PPP: NOMÉS total de finca (no per unitat, no és qualificació especial) ---
        try:
            lyr = None
            ja_retallada = False
            if gpkg_path:
                uri = gpkg_path + "|layername=PPP"
                cand = QgsVectorLayer(uri, "PPP", "ogr")
                if cand.isValid():
                    lyr, ja_retallada = cand, True
            if lyr is None:
                shp_path = _descarrega_shp_zip(URL_PPP, "ppp")
                lyr = QgsVectorLayer(shp_path, "ppp_tmp", "ogr")
            if lyr.isValid():
                import processing
                res_clip = processing.run("native:clip", {
                    "INPUT": lyr, "OVERLAY": ambit_lyr, "OUTPUT": "TEMPORARY_OUTPUT",
                })
                clipped = res_clip["OUTPUT"]
                camp_nom_ppp = _troba_camp_output(clipped, "NOM")
                for feat in clipped.getFeatures():
                    geom = feat.geometry()
                    if geom and not geom.isEmpty():
                        area_ha = geom.area() / 10000.0
                        ppp_total_finca += area_ha
                        nom_ppp = None
                        if camp_nom_ppp and camp_nom_ppp in feat.fields().names():
                            nom_ppp = feat[camp_nom_ppp]
                        nom_ppp = (nom_ppp or "").strip() or "(sense nom)"
                        ppp_per_nom[nom_ppp] = ppp_per_nom.get(nom_ppp, 0.0) + area_ha
                consultades_ok.append("PPP")
            else:
                errors.append("PPP")
        except Exception:
            errors.append("PPP")

        # --- Tipus de risc d'incendis (finca sencera) ---
        tipus_risc_incendi = None
        try:
            valors_risc = _troba_risc_incendi_tipus(ambit_lyr)
            if valors_risc:
                tipus_risc_incendi = " / ".join(sorted(valors_risc))
            consultades_ok.append("Tipus de risc d'incendis")
        except Exception as e:
            _log_error("Tipus de risc d'incendis", e)
            errors.append("Tipus de risc d'incendis")

        # Índex de perill: mateix valor que el tipus de risc d'incendis
        # (l'usuari ha demanat treure la font INFOCAT -- ja no es
        # consulta cap servei nou, es reutilitza directament aquest
        # valor).
        index_perill = tipus_risc_incendi

        # --- 8. LU: mostreig per punt (punt garantit dins de cada unitat) ---
        try:
            lu_total = 0
            lu_fallides = 0
            for feat in unitats_lyr.getFeatures():
                geom = feat.geometry()
                if not geom or geom.isEmpty():
                    continue
                # pointOnSurface(), no centroid(): el centroide d'una
                # unitat molt allargada o còncava pot caure fora del
                # propi polígon, fent que la consulta WMS mostregi un
                # punt equivocat (d'una altra unitat o fora de l'IOF).
                punt_mostreig = geom.pointOnSurface().asPoint()
                lu_total += 1
                try:
                    codi_muc, desc_ajunt = _get_feature_info_muc(punt_mostreig.x(), punt_mostreig.y())
                except Exception as e:
                    lu_fallides += 1
                    _log_error("LU (GetFeatureInfo MUC)", e)
                    continue
                if codi_muc and codi_muc.strip().upper().startswith("N2"):
                    area_ha = geom.area() / 10000.0
                    codi = str(feat[codi_field]) if codi_field else ""
                    nom_lu = (desc_ajunt or "").strip() or "(sense nom -- citar el municipi a Observacions)"
                    resum_total[("LU", nom_lu)] = resum_total.get(("LU", nom_lu), 0.0) + area_ha
                    key = (codi, "LU", nom_lu)
                    resum_per_unitat[key] = resum_per_unitat.get(key, 0.0) + area_ha
                    geometries_per_unitat.setdefault(codi, []).append(QgsGeometry(geom))
            if lu_fallides == 0:
                consultades_ok.append("LU")
            elif lu_fallides < lu_total:
                errors.append(
                    "LU (" + str(lu_fallides) + " de " + str(lu_total) +
                    " unitats no consultades -- revisa la connexió)"
                )
            else:
                errors.append("LU (servei no disponible -- cap unitat s'ha pogut consultar)")
        except Exception:
            errors.append("LU")

    finally:
        QApplication.restoreOverrideCursor()

    if not _confirma_abans_exportar(win, consultades_ok, errors):
        return

    if not resum_total and ppp_total_finca <= 0:
        text = "No s'ha trobat cap afectació dins l'àmbit de l'IOF."
        if errors:
            text += ("\n\nA més, no s'han pogut consultar aquestes "
                     "qualificacions:\n" + "\n".join("  • " + e for e in errors))
        QMessageBox.information(win, "Qualificacions especials", text)
        return

    # Superfície REAL afectada per unitat (unió de totes les geometries de
    # qualificacions que la toquen, no la suma -- si no, una zona que és
    # alhora ENPE i PEIN comptaria dues vegades la mateixa superfície).
    # Es valida que mai sigui superior a la superfície total ordenada de
    # la unitat (ni la de tot l'IOF).
    superficie_afectada_per_unitat = {}
    superficie_total_per_unitat = {}
    totes_les_geometries = []
    for feat in unitats_lyr.getFeatures():
        geom = feat.geometry()
        if not geom or geom.isEmpty():
            continue
        codi = str(feat[codi_field]) if codi_field else ""
        superficie_total_per_unitat[codi] = superficie_total_per_unitat.get(codi, 0.0) + geom.area() / 10000.0

    for codi, geoms in geometries_per_unitat.items():
        area_unio = _calcula_superficie_unio(geoms)
        area_total_unitat = superficie_total_per_unitat.get(codi, 0.0)
        # Marge de tolerància per precisió numèrica (topologia de fonts diferents)
        if area_total_unitat > 0 and area_unio > area_total_unitat * 1.001:
            area_unio = area_total_unitat
        superficie_afectada_per_unitat[codi] = area_unio
        totes_les_geometries.extend(geoms)

    superficie_total_iof = sum(superficie_total_per_unitat.values())
    superficie_afectada_iof = _calcula_superficie_unio(totes_les_geometries)
    if superficie_total_iof > 0 and superficie_afectada_iof > superficie_total_iof * 1.001:
        superficie_afectada_iof = superficie_total_iof

    # Detecció defensiva d'una inconsistència possible: si hi ha
    # qualificacions amb superfície real trobada (resum_total no buit)
    # però la unió surt a 0 (normalment perquè geometries_per_unitat ha
    # quedat buit, p. ex. si les tipologies forestals no s'han definit
    # correctament o no coincideixen geomètricament amb les
    # interseccions), s'avisa explícitament -- en lloc de mostrar
    # silenciosament un "0" que sembla dir que no hi ha cap afectació.
    suma_resum_total = sum(resum_total.values())
    if suma_resum_total > 0 and superficie_afectada_iof <= 0:
        QMessageBox.warning(
            win, "Superfície afectada no calculable",
            "S'han trobat qualificacions especials amb superfície real "
            "(" + _format_ha_ca(suma_resum_total) + " ha en total, sumant "
            "totes les qualificacions), però no s'ha pogut calcular la "
            "superfície afectada per unitat ni per al conjunt de l'IOF.\n\n"
            "Això sol passar quan les tipologies forestals no estan "
            "definides correctament (per exemple, si no coincideixen "
            "geomètricament amb la resta de capes). Revisa que les "
            "tipologies forestals estiguin ben digitalitzades abans de "
            "confiar en els valors de \"Superfície afectada\" de "
            "l'informe -- la resta de dades (superfície per "
            "qualificació) sí és fiable."
        )

    # Cerca de correspondència amb el PDF per a cada (qualificacio, nom) trobat
    correspondencies = {}
    for qualif, nom in noms_trobats:
        if qualif in ("FAUNA", "LU"):
            continue  # aquests no es cerquen per nom individual
        if qualif == "ZAU":
            # El propi CODI_ZAU ja és el text del PDF (coincidencia ~100%)
            correspondencies[(qualif, nom)] = [nom]
            continue
        if qualif == "UP":
            correspondencies[(qualif, nom)] = ["BOSC PROTECTOR"]
            continue
        correspondencies[(qualif, nom)] = _cerca_correspondencia_pdf(qualif, nom)

    fitxer = _genera_informe(
        win, resum_total, resum_per_unitat, codi_field,
        correspondencies, fauna_especies, ppp_total_finca,
        superficie_afectada_per_unitat, superficie_total_per_unitat,
        superficie_afectada_iof, superficie_total_iof,
        tipus_risc_incendi, index_perill, ppp_per_nom,
    )

    text = "Superfície total afectada per qualificacions especials:\n\n"
    per_qualif = {}
    for (qualif, nom), area in resum_total.items():
        per_qualif[qualif] = per_qualif.get(qualif, 0.0) + area
    for qualif, area in sorted(per_qualif.items(), key=lambda x: -x[1]):
        text += "  • " + qualif + ": " + _format_ha_ca(area) + " ha\n"
    if fauna_especies:
        text += "\nEspècies de FAUNA trobades (citar a Observacions):\n"
        text += "\n".join("  • " + e for e in sorted(fauna_especies)) + "\n"
    text += "\nLa cartografia de Boscos Singulars no està disponible.\n"
    text += "\nSuperfície total inclosa en PPP: " + _format_ha_ca(ppp_total_finca) + " ha\n"
    if errors:
        text += ("\nNo s'han pogut consultar (revisar connexió a internet):\n" +
                  "\n".join("  • " + e for e in errors) + "\n")
    if fitxer:
        text += "\nInforme desat a:\n" + fitxer
    QMessageBox.information(win, "Qualificacions especials calculades", text)


def _filtra_per_expressio(layer, expressio):
    """Retorna una còpia filtrada de la capa segons l'expressió donada
    (native:extractbyexpression), o None si falla."""
    try:
        import processing
        res = processing.run("native:extractbyexpression", {
            "INPUT": layer, "EXPRESSION": expressio, "OUTPUT": "TEMPORARY_OUTPUT",
        })
        return res["OUTPUT"]
    except Exception:
        return None


def _genera_informe(win, resum_total, resum_per_unitat, codi_field,
                     correspondencies, fauna_especies, ppp_total_finca,
                     superficie_afectada_per_unitat=None,
                     superficie_total_per_unitat=None,
                     superficie_afectada_iof=None,
                     superficie_total_iof=None,
                     tipus_risc_incendi=None, index_perill=None,
                     ppp_per_nom=None):
    """Genera l'informe (Excel si openpyxl és disponible, si no text pla),
    demanant a l'usuari on desar-lo. Indica per a cada espai si té
    correspondència amb el desplegable del formulari PDF, o si cal
    citar-lo manualment a Observacions. Retorna el path del fitxer, o
    None si s'ha cancel·lat.

    superficie_afectada_per_unitat / superficie_total_per_unitat: la
    superfície afectada és la UNIÓ real de totes les qualificacions que
    toquen la unitat (mai la suma -- una zona ENPE i PEIN alhora no pot
    comptar's dues vegades), sempre <= la superfície total de la unitat.
    Igual pel conjunt de l'IOF (superficie_afectada_iof / _total_iof).

    tipus_risc_incendi / index_perill: pendents d'integrar (fonts
    d'incendis forestals encara en investigació) -- si són None es
    mostren com "Pendent".
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        te_excel = True
    except ImportError:
        te_excel = False

    filtre = "Excel (*.xlsx)" if te_excel else "Text (*.txt)"
    ext_defecte = ".xlsx" if te_excel else ".txt"
    nom_iof = _troba_nom_iof()
    nom_fitxer = (
        "IOF_" + nom_iof + "-Qualificacions_especials"
        if nom_iof else "IOF_Qualificacions_especials"
    )
    # Sanejament bàsic: caràcters no vàlids en noms de fitxer (Windows)
    nom_fitxer = re.sub(r'[\\/:*?"<>|]', "_", nom_fitxer)
    path, _ = QFileDialog.getSaveFileName(
        win, "Desar informe de qualificacions especials",
        nom_fitxer + ext_defecte, filtre
    )
    if not path:
        return None

    def _text_correspondencia(qualif, nom):
        if qualif == "FAUNA":
            return ""
        if qualif == "LU":
            return "SENSE CORRESPONDÈNCIA"
        pdf_list = correspondencies.get((qualif, nom), [])
        if pdf_list:
            return " / ".join(pdf_list)
        return "SENSE CORRESPONDÈNCIA"

    def _nom_qualif(qualif):
        return NOMS_QUALIFICACIO_INFORME.get(qualif, qualif)

    COLORS_QUALIF = {
        "ENPE": "C6E0B4", "PEIN": "C6E0B4", "PEIN-PE": "C6E0B4", "RF": "C6E0B4",
        "LIC": "BDD7EE", "ZEPA": "BDD7EE", "LIC-ZEPA": "BDD7EE",
        "UP": "FFE699", "FAUNA": "FFD966", "ZAU": "D9D2E9", "LU": "F4CCCC",
    }

    if te_excel and path.lower().endswith(".xlsx"):
        FONT_TITOL = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        FONT_CAPCALERA = Font(name="Arial", size=10, bold=True)
        FONT_NORMAL = Font(name="Arial", size=10)
        FILL_TITOL = PatternFill("solid", fgColor="4472C4")
        FILL_CAPCALERA = PatternFill("solid", fgColor="D9D9D9")
        VORA_FINA = Side(style="thin", color="BFBFBF")
        VORA_CEL = Border(left=VORA_FINA, right=VORA_FINA, top=VORA_FINA, bottom=VORA_FINA)
        FILL_FILA_ALTERNA = PatternFill("solid", fgColor="F2F2F2")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Totals per finca"

        # --- Títol "Qualificacions especials" ---
        ws.merge_cells("A1:D1")
        ws["A1"] = "Qualificacions especials"
        ws["A1"].font = FONT_TITOL
        ws["A1"].fill = FILL_TITOL
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        capçaleres = ["Qualificació", "Nom de l'espai", "Superfície (ha)", "Correspondència amb el PDF"]
        for col, text in enumerate(capçaleres, start=1):
            c = ws.cell(row=2, column=col, value=text)
            c.font = FONT_CAPCALERA
            c.fill = FILL_CAPCALERA

        row = 3
        for (qualif, nom), area in sorted(resum_total.items(), key=lambda x: -x[1]):
            fill_color = COLORS_QUALIF.get(qualif)
            fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
            valors = [_nom_qualif(qualif), nom, round(area, 2), _text_correspondencia(qualif, nom)]
            for col, v in enumerate(valors, start=1):
                cel = ws.cell(row=row, column=col, value=v)
                cel.font = FONT_NORMAL
                if fill:
                    cel.fill = fill
            row += 1

        row += 1

        if fauna_especies:
            row += 1
            ws.cell(row=row, column=1, value="Espècies de FAUNA (citar a Observacions)").font = FONT_CAPCALERA
            row += 1
            FONT_CURSIVA = Font(name="Arial", size=10, italic=True)
            for esp in sorted(fauna_especies):
                ws.cell(row=row, column=2, value=esp).font = FONT_CURSIVA
                row += 1

        row += 1
        ws.cell(row=row, column=1, value="La cartografia de Boscos Singulars no està disponible.").font = FONT_NORMAL
        row += 1

        if superficie_afectada_iof is not None:
            row += 1
            ws.cell(row=row, column=1, value="Superfície total ordenada de l'IOF (ha)").font = FONT_CAPCALERA
            ws.cell(row=row, column=3, value=round(superficie_total_iof, 2)).font = FONT_NORMAL
            row += 1
            ws.cell(row=row, column=1, value="Superfície afectada per qualificacions especials (ha)").font = FONT_CAPCALERA
            ws.cell(row=row, column=3, value=round(superficie_afectada_iof, 2)).font = FONT_NORMAL
            row += 1

        # --- Títol "Incendis forestal", deixant una línia lliure ---
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value="Incendis forestal")
        ws.cell(row=row, column=1).font = FONT_TITOL
        ws.cell(row=row, column=1).fill = FILL_TITOL
        ws.row_dimensions[row].height = 24
        row += 1
        for col, text in enumerate(["PPP", "Nom de l'espai", "Superfície (ha)"], start=1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = FONT_CAPCALERA
            c.fill = FILL_CAPCALERA
        row += 1
        if ppp_per_nom:
            for nom_ppp, area_ppp in sorted(ppp_per_nom.items(), key=lambda x: -x[1]):
                ws.cell(row=row, column=2, value=nom_ppp).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=round(area_ppp, 2)).font = FONT_NORMAL
                row += 1
        else:
            ws.cell(row=row, column=3, value=round(ppp_total_finca, 2)).font = FONT_NORMAL
            row += 1
        ws.cell(row=row, column=2, value="Total").font = FONT_CAPCALERA
        ws.cell(row=row, column=3, value=round(ppp_total_finca, 2)).font = FONT_CAPCALERA
        row += 2

        for col, text in enumerate(["Tipus de risc d'incendis", "Índex de perill"], start=1):
            c = ws.cell(row=row, column=col, value=text)
            c.font = FONT_CAPCALERA
            c.fill = FILL_CAPCALERA
        row += 1
        ws.cell(row=row, column=1, value=tipus_risc_incendi if tipus_risc_incendi else "Pendent").font = FONT_NORMAL
        ws.cell(row=row, column=2, value=index_perill if index_perill else "Pendent").font = FONT_NORMAL

        for col, ample in [(1, 40), (2, 45), (3, 16), (4, 55)]:
            ws.column_dimensions[get_column_letter(col)].width = ample
        ws.freeze_panes = "A3"

        # --- Full "Per unitat": només la superfície total afectada, per unitat ---
        ws_unitat = wb.create_sheet("Per unitat")
        ws_unitat.merge_cells("A1:D1")
        ws_unitat["A1"] = "Superfície afectada per unitat d'actuació"
        ws_unitat["A1"].font = FONT_TITOL
        ws_unitat["A1"].fill = FILL_TITOL
        ws_unitat.row_dimensions[1].height = 24
        for col, text in enumerate(
            ["Unitat d'actuació", "Superfície total ordenada (ha)",
             "Superfície afectada per qualificacions especials (ha)", "% afectat"],
            start=1,
        ):
            c = ws_unitat.cell(row=2, column=col, value=text)
            c.font = FONT_CAPCALERA
            c.fill = FILL_CAPCALERA
            c.border = VORA_CEL

        codis = sorted(set(list(superficie_total_per_unitat or {}) + list(superficie_afectada_per_unitat or {})))
        row = 3
        for i, codi in enumerate(codis):
            total_u = (superficie_total_per_unitat or {}).get(codi, 0.0)
            afectada_u = (superficie_afectada_per_unitat or {}).get(codi, 0.0)
            pct = (afectada_u / total_u * 100) if total_u > 0 else 0.0
            fill_fila = FILL_FILA_ALTERNA if i % 2 == 1 else None
            for col, v in enumerate([codi, round(total_u, 2), round(afectada_u, 2), round(pct, 1)], start=1):
                cel = ws_unitat.cell(row=row, column=col, value=v)
                cel.font = FONT_NORMAL
                cel.border = VORA_CEL
                if fill_fila:
                    cel.fill = fill_fila
            row += 1
        for col, ample in [(1, 20), (2, 26), (3, 40), (4, 12)]:
            ws_unitat.column_dimensions[get_column_letter(col)].width = ample
        ws_unitat.freeze_panes = "A3"

        wb.save(path)
    else:
        with open(path, "w", encoding="utf-8") as f:
            f.write("QUALIFICACIONS ESPECIALS\n")
            f.write("=" * 60 + "\n")
            for (qualif, nom), area in sorted(resum_total.items(), key=lambda x: -x[1]):
                f.write(_nom_qualif(qualif) + " -- " + nom + ": " + format(area, ".2f") + " ha")
                corr = _text_correspondencia(qualif, nom)
                if corr:
                    f.write(" [" + corr + "]")
                f.write("\n")
            if fauna_especies:
                f.write("\nEspècies de FAUNA (citar a Observacions):\n")
                for esp in sorted(fauna_especies):
                    f.write("  - *" + esp + "*\n")
            f.write("\nLa cartografia de Boscos Singulars no està disponible.\n")
            if superficie_afectada_iof is not None:
                f.write("\nSuperfície total ordenada de l'IOF: " + format(superficie_total_iof, ".2f") + " ha\n")
                f.write("Superfície afectada per qualificacions especials: " +
                        format(superficie_afectada_iof, ".2f") + " ha\n")

            f.write("\nINCENDIS FORESTAL\n")
            f.write("=" * 60 + "\n")
            f.write("Tipus de risc d'incendis: " + (tipus_risc_incendi or "Pendent") + "\n")
            f.write("Índex de perill: " + (index_perill or "Pendent") + "\n")
            f.write("\nPPP:\n")
            if ppp_per_nom:
                for nom_ppp, area_ppp in sorted(ppp_per_nom.items(), key=lambda x: -x[1]):
                    f.write("  - " + nom_ppp + ": " + format(area_ppp, ".2f") + " ha\n")
            f.write("  Total PPP: " + format(ppp_total_finca, ".2f") + " ha\n")

            f.write("\nSUPERFÍCIE AFECTADA PER UNITAT D'ACTUACIÓ\n")
            f.write("=" * 60 + "\n")
            codis = sorted(set(list(superficie_total_per_unitat or {}) + list(superficie_afectada_per_unitat or {})))
            for codi in codis:
                total_u = (superficie_total_per_unitat or {}).get(codi, 0.0)
                afectada_u = (superficie_afectada_per_unitat or {}).get(codi, 0.0)
                f.write("UA " + codi + ": " + format(afectada_u, ".2f") + " ha afectades de " +
                        format(total_u, ".2f") + " ha totals\n")
    return path

